"""Application service for document fallback and template workbook generation."""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
import re
import tempfile
import time
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import load_workbook
from pypdf.errors import PyPdfError

import estonia_extractor as extractor
from api_adapter import (
    fallback_document_candidates,
    merge_missing,
    needs_document_fallback,
    replace_unconsolidated_with_consolidated,
    select_annual_report_pdf,
    select_fallback_document,
    select_xbrl_document,
)
from rik_xml_client import CompanyDocument, DownloadedDocument, RikError, RikXmlClient
from workbook_preservation import restore_extended_validations
from xbrl_parser import XbrlParseError, parse_xbrl_report


@dataclass(frozen=True)
class GeneratedWorkbook:
    content: bytes
    filename: str
    messages: list[str]
    source_details: list[str]


@dataclass(frozen=True)
class AnnualReportPdfBundle:
    content: bytes
    filename: str
    included_years: tuple[int, ...]
    warnings: tuple[str, ...]


def safe_company_name(value: str) -> str:
    value = re.sub(r'[<>:"/\\|?*]+', "", value or "").strip()
    value = re.sub(r"\s+", " ", value)
    return value[:120] or "Estonia Company"


def build_annual_report_pdf_bundle(
    client: RikXmlClient,
    documents: list[CompanyDocument],
    fiscal_years: list[int],
    *,
    company_name: str,
    cached_documents: dict[int, DownloadedDocument] | None = None,
) -> AnnualReportPdfBundle:
    selected_years = sorted(set(fiscal_years), reverse=True)
    cached_documents = cached_documents or {}
    downloaded_by_year: dict[int, DownloadedDocument] = {}
    warnings: list[str] = []
    for year in selected_years:
        document = select_annual_report_pdf(documents, year)
        if document is None:
            warnings.append(f"FY{year}: complete annual-report PDF was not available from RIK.")
            continue
        try:
            downloaded = cached_documents.get(year) or client.download_document(document)
        except RikError as exc:
            warnings.append(f"FY{year}: complete annual-report PDF could not be downloaded ({exc}).")
            continue
        if b"%PDF-" not in downloaded.content[:1024]:
            warnings.append(f"FY{year}: RIK's annual-report download was not a valid PDF.")
            continue
        downloaded_by_year[year] = downloaded

    output = BytesIO()
    if downloaded_by_year:
        with ZipFile(output, "w", compression=ZIP_DEFLATED) as archive:
            for year in sorted(downloaded_by_year, reverse=True):
                archive.writestr(
                    f"{safe_company_name(company_name)} AR {year}.pdf",
                    downloaded_by_year[year].content,
                )
    year_label = (
        str(selected_years[0])
        if len(selected_years) == 1
        else f"{selected_years[-1]}-{selected_years[0]}"
        if selected_years
        else "Annual Reports"
    )
    return AnnualReportPdfBundle(
        content=output.getvalue(),
        filename=f"{safe_company_name(company_name)} Annual Reports {year_label}.zip",
        included_years=tuple(sorted(downloaded_by_year, reverse=True)),
        warnings=tuple(warnings),
    )


def workbook_source_details(
    jobs: list[extractor.FillJob],
    *,
    fill_segments: bool,
    fill_cogs: bool,
    fill_goodwill_amortisation: bool,
) -> list[str]:
    enabled_flags = {
        "fill_cogs": fill_cogs,
        "fill_goodwill_amortisation": fill_goodwill_amortisation,
    }
    details: list[str] = []
    for job in sorted(jobs, key=lambda item: item.year, reverse=True):
        source_year = job.source_year or job.report.year
        source_period = (
            "comparative"
            if source_year == job.year + 1
            else "current"
            if source_year == job.year
            else "available figures"
        )
        written_items: list[str] = []
        for mapping in extractor.WORKBOOK_MAPPINGS:
            if mapping.optional_flag and not enabled_flags.get(mapping.optional_flag, False):
                continue
            item = mapping.item
            raw_value = job.report.get_value(job.value_year, item)
            if item in {"Total depreciation", "Total amortisation"} and job.report.get_value(
                job.value_year, "D&A"
            ) is not None:
                continue
            if item in {"Investments in tangible assets", "Investments in intangible assets"} and job.report.get_value(
                job.value_year, "CAPEX"
            ) is not None:
                continue
            if item == "D&A" and raw_value is None and (
                job.report.get_value(job.value_year, "Total depreciation") is not None
                or job.report.get_value(job.value_year, "Total amortisation") is not None
            ):
                continue
            if item == "CAPEX" and raw_value is None and (
                job.report.get_value(job.value_year, "Investments in tangible assets") is not None
                or job.report.get_value(job.value_year, "Investments in intangible assets") is not None
            ):
                continue
            if raw_value is not None and item not in written_items:
                written_items.append(item)
        xbrl_items = [
            item
            for item in written_items
            if (job.report.get_source(job.value_year, item) or "").startswith("RIK XBRL |")
        ]
        xml_items = [
            item
            for item in written_items
            if (job.report.get_source(job.value_year, item) or "").startswith("RIK XML |")
        ]
        fallback_items = [
            item for item in written_items if item not in xbrl_items and item not in xml_items
        ]
        parts = [f"FY{job.year} ← AR{source_year} {source_period}"]
        if xbrl_items:
            xbrl_label = (
                "consolidated XBRL items"
                if job.report.accounting_basis == "consolidated"
                else "XBRL items"
            )
            parts.append(f"{xbrl_label}: {', '.join(xbrl_items)}")
        if xml_items:
            parts.append(f"statement XML items: {', '.join(xml_items)}")
        if fallback_items:
            parts.append(f"PDF/BDOC fallback items: {', '.join(fallback_items)}")
        if not (xbrl_items or xml_items or fallback_items):
            parts.append("no mapped items")
        if fill_segments:
            segment_counts: dict[str, list[str]] = {"XBRL": [], "PDF/BDOC fallback": []}
            for segment_by, records in sorted(job.report.segments.get(job.value_year, {}).items()):
                xbrl_count = sum(
                    (record.source or "").startswith("RIK XBRL |")
                    for record in records.values()
                )
                fallback_count = len(records) - xbrl_count
                if xbrl_count:
                    segment_counts["XBRL"].append(f"{segment_by} ({xbrl_count})")
                if fallback_count:
                    segment_counts["PDF/BDOC fallback"].append(
                        f"{segment_by} ({fallback_count})"
                    )
            for label, counts in segment_counts.items():
                if counts:
                    parts.append(f"{label} segmentations: {', '.join(counts)}")
        details.append(" — ".join(parts) + ".")
    return details


def add_xbrl_sources(
    client: RikXmlClient,
    reports: list[extractor.EstonianReport],
    documents: list[CompanyDocument],
    fiscal_years: list[int],
    *,
    registry_code: str,
    company_name: str,
    retry_delay_seconds: float = 16.0,
) -> tuple[list[extractor.EstonianReport], list[str]]:
    """Make scope-matched annual-report XBRL the primary structured source."""
    reports_by_year = {report.year: report for report in reports if report.period_end is not None}
    warnings: list[str] = []
    for year in sorted(set(fiscal_years), reverse=True):
        document = select_xbrl_document(documents, year)
        if document is None:
            warnings.append(f"FY{year}: no valid annual-report XBRL package was available from RIK.")
            continue
        parsed = None
        errors: list[str] = []
        for attempt in range(2):
            try:
                downloaded = client.download_document(document)
                parsed = parse_xbrl_report(
                    downloaded.content,
                    registry_code=registry_code,
                    company_name=company_name,
                    report_year=year,
                    source_name=downloaded.filename,
                )
                break
            except (RikError, XbrlParseError) as exc:
                errors.append(str(exc))
                if attempt == 0 and retry_delay_seconds > 0:
                    time.sleep(retry_delay_seconds)
        if parsed is None:
            warnings.append(
                f"FY{year}: XBRL could not be used ({errors[-1] if errors else 'unknown error'}); "
                "statement XML/PDF fallback remains active."
            )
            continue

        prior = reports_by_year.get(year)
        if prior is not None:
            if prior.accounting_basis == "consolidated" and parsed.accounting_basis != "consolidated":
                merge_missing(prior, parsed)
                warnings.append(
                    f"FY{year}: the XBRL package lacked consolidated primary statements, so the "
                    "consolidated statement XML remained primary."
                )
                continue
            if not (
                parsed.accounting_basis == "consolidated"
                and prior.accounting_basis == "unconsolidated"
            ):
                merge_missing(parsed, prior)
            for index, existing in enumerate(reports):
                if existing is prior:
                    reports[index] = parsed
                    break
        else:
            reports.append(parsed)
        reports_by_year[year] = parsed
    return reports, warnings


def add_document_fallbacks(
    client: RikXmlClient,
    reports: list[extractor.EstonianReport],
    documents: list[CompanyDocument],
    fiscal_years: list[int],
    *,
    company_name: str,
    downloaded_documents: dict[int, DownloadedDocument] | None = None,
) -> tuple[list[extractor.EstonianReport], list[str]]:
    reports_by_year = {report.year: report for report in reports if report.period_end is not None}
    warnings: list[str] = []
    with tempfile.TemporaryDirectory() as temporary_directory:
        temporary_path = Path(temporary_directory)
        for year in sorted(set(fiscal_years), reverse=True):
            primary = reports_by_year.get(year)
            if primary is not None and not needs_document_fallback(primary, year):
                continue
            candidates = fallback_document_candidates(documents, year)
            if not candidates:
                warnings.append(f"FY{year}: no PDF or BDOC fallback document was available.")
                continue
            fallback = None
            errors: list[str] = []
            for document in candidates:
                try:
                    downloaded = client.download_document(document)
                    suffix = document.extension or Path(downloaded.filename).suffix or ".bin"
                    source_path = temporary_path / f"{document.document_id}_{year}{suffix}"
                    source_path.write_bytes(downloaded.content)
                    parsed_reports = extractor.parse_reports_from_input(
                        source_path, extractor.load_terms(None)
                    )
                    fallback = next(
                        (candidate for candidate in parsed_reports if candidate.year == year),
                        parsed_reports[0],
                    )
                    fallback.source_path = Path(downloaded.filename)
                    if not fallback.company:
                        fallback.company = company_name
                    if downloaded_documents is not None and document.document_type == "A":
                        downloaded_documents[year] = downloaded
                    break
                except (RikError, OSError, ValueError, PyPdfError, IndexError) as exc:
                    errors.append(f"{document.document_type}: {exc}")
            if fallback is None:
                warnings.append(
                    f"FY{year}: document fallback could not be parsed "
                    f"({'; '.join(errors) or 'no readable annual report'})."
                )
                continue
            if primary is None:
                reports.append(fallback)
                reports_by_year[year] = fallback
            else:
                replaced = replace_unconsolidated_with_consolidated(primary, fallback)
                if replaced:
                    replaced_items = sorted({item for items in replaced.values() for item in items})
                    warnings.append(
                        f"FY{year}: consolidated annual-report figures replaced explicitly "
                        f"unconsolidated XML for {', '.join(replaced_items)}."
                    )
                else:
                    merge_missing(primary, fallback)
    return reports, warnings


def generate_workbook(
    reports: list[extractor.EstonianReport],
    fiscal_years: list[int],
    template_path: Path,
    output_path: Path,
    *,
    company_name: str,
    fill_segments: bool = True,
    fill_cogs: bool = True,
    fill_goodwill_amortisation: bool = False,
) -> GeneratedWorkbook:
    if not reports:
        raise ValueError("No structured or document report data was available to fill the workbook.")
    selected_years = sorted(set(fiscal_years), reverse=True)
    jobs, messages = extractor.build_fill_jobs(
        reports,
        years=max(len(selected_years), 1),
        fill_comparative=True,
        target_years=selected_years,
    )
    jobs = [job for job in jobs if job.year in selected_years]
    if not jobs:
        raise ValueError("None of the selected fiscal years could be prepared for the workbook.")

    workbook = load_workbook(template_path)
    if extractor.FINANCIALS_SHEET not in workbook.sheetnames:
        raise ValueError("The embedded template does not contain the Financials sheet.")
    worksheet = workbook[extractor.FINANCIALS_SHEET]

    for job in jobs:
        messages.append(
            f"=== Filling FY{job.year} from {job.report.payload_name} values FY{job.value_year} ==="
        )
        messages.extend(
            extractor.fill_period(
                worksheet,
                job,
                overwrite=True,
                add_comments=True,
                fill_cogs=fill_cogs,
                fill_goodwill_amortisation=fill_goodwill_amortisation,
                dry_run=False,
            )
        )
    messages.extend(extractor.update_cagr_formulas(worksheet, jobs, dry_run=False))
    messages.extend(extractor.group_unused_year_columns(worksheet, jobs, dry_run=False))
    if fill_segments:
        messages.extend(
            extractor.fill_segments_sheet(
                workbook,
                jobs,
                overwrite=True,
                add_comments=True,
                dry_run=False,
            )
        )

    extractor.save_workbook(workbook, output_path)
    restore_extended_validations(template_path, output_path)
    filename = f"{safe_company_name(company_name)} Financials.xlsx"
    source_details = workbook_source_details(
        jobs,
        fill_segments=fill_segments,
        fill_cogs=fill_cogs,
        fill_goodwill_amortisation=fill_goodwill_amortisation,
    )
    return GeneratedWorkbook(
        content=output_path.read_bytes(),
        filename=filename,
        messages=messages,
        source_details=source_details,
    )
