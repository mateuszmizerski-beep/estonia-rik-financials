from __future__ import annotations

from datetime import date, datetime, timezone
from decimal import Decimal
from io import BytesIO
import json
from pathlib import Path
import tempfile
import unittest
from zipfile import ZipFile

from openpyxl import load_workbook

from api_adapter import (
    fallback_document_candidates,
    merge_missing,
    preferred_report_types,
    replace_unconsolidated_with_consolidated,
    report_from_lines,
    source_report_years,
)
from estonia_extractor import (
    EstonianReport,
    SegmentRecord,
    build_fill_jobs,
    build_segment_summary_rows,
    extract_financial_values,
    load_terms,
)
from financials_service import (
    add_xbrl_sources,
    build_annual_report_pdf_bundle,
    generate_workbook,
    workbook_source_details,
)
from rik_xml_client import (
    AnnualReportAvailability,
    CompanyDocument,
    DownloadedDocument,
    StatementLine,
)
from workbook_preservation import count_extended_validations
from xbrl_parser import XbrlParseError, parse_xbrl_report


PROJECT = Path(__file__).parents[1]
TEMPLATE = PROJECT / "assets" / "gainpro_template_eur.xlsx"
LIVE_FIXTURES = PROJECT / "tests" / "fixtures" / "live"


def availability(report_type: str, year: int = 2024) -> AnnualReportAvailability:
    return AnnualReportAvailability(
        report_type,
        f"Report {report_type}",
        year,
        date(year, 1, 1),
        date(year, 12, 31),
    )


def line(
    report_type: str,
    row_number: str,
    row_name: str,
    column: str,
    value: str,
) -> StatementLine:
    return StatementLine(
        registry_code="70000310",
        fiscal_year=2024,
        report_type=report_type,
        report_name=f"Report {report_type}",
        period_start=date(2024, 1, 1),
        period_end=date(2024, 12, 31),
        row_number=row_number,
        row_name=row_name,
        column_code=column,
        column_name="Current" if column == "A1" else "Previous",
        value=Decimal(value),
        retrieved_at=datetime(2025, 1, 2, tzinfo=timezone.utc),
    )


def fixture_lines(registry_code: str, year: int) -> list[StatementLine]:
    payload = json.loads((LIVE_FIXTURES / f"{registry_code}.json").read_text(encoding="utf-8"))
    lines: list[StatementLine] = []
    for statement in payload["statements"]:
        if statement["fiscal_year"] != year:
            continue
        for item in statement["lines"]:
            lines.append(
                StatementLine(
                    registry_code=item["registry_code"],
                    fiscal_year=item["fiscal_year"],
                    report_type=item["report_type"],
                    report_name=item["report_name"],
                    period_start=date.fromisoformat(item["period_start"])
                    if item["period_start"]
                    else None,
                    period_end=date.fromisoformat(item["period_end"])
                    if item["period_end"]
                    else None,
                    row_number=item["row_number"],
                    row_name=item["row_name"],
                    column_code=item["column_code"],
                    column_name=item["column_name"],
                    value=Decimal(item["value"]) if item["value"] is not None else None,
                    retrieved_at=datetime.fromisoformat(item["retrieved_at"]),
                )
            )
    return lines


def consolidated_xbrl_package() -> bytes:
    xml = b'''<?xml version="1.0" encoding="UTF-8"?>
<xbrli:xbrl xmlns:xbrli="http://www.xbrl.org/2003/instance"
 xmlns:xbrldi="http://xbrl.org/2006/xbrldi" xmlns:et-gaap="urn:et-gaap">
 <xbrli:context id="I1"><xbrli:entity><xbrli:identifier scheme="urn:test">1</xbrli:identifier></xbrli:entity><xbrli:period><xbrli:instant>2025-12-31</xbrli:instant></xbrli:period></xbrli:context>
 <xbrli:context id="I2"><xbrli:entity><xbrli:identifier scheme="urn:test">1</xbrli:identifier></xbrli:entity><xbrli:period><xbrli:instant>2024-12-31</xbrli:instant></xbrli:period></xbrli:context>
 <xbrli:context id="D1"><xbrli:entity><xbrli:identifier scheme="urn:test">1</xbrli:identifier></xbrli:entity><xbrli:period><xbrli:startDate>2025-01-01</xbrli:startDate><xbrli:endDate>2025-12-31</xbrli:endDate></xbrli:period></xbrli:context>
 <xbrli:context id="D2"><xbrli:entity><xbrli:identifier scheme="urn:test">1</xbrli:identifier></xbrli:entity><xbrli:period><xbrli:startDate>2024-01-01</xbrli:startDate><xbrli:endDate>2024-12-31</xbrli:endDate></xbrli:period></xbrli:context>
 <xbrli:context id="TR1"><xbrli:entity><xbrli:identifier scheme="urn:test">1</xbrli:identifier><xbrli:segment><xbrldi:explicitMember dimension="et-gaap:MaturityDimension">et-gaap:MaturityTotalAbstract</xbrldi:explicitMember></xbrli:segment></xbrli:entity><xbrli:period><xbrli:instant>2025-12-31</xbrli:instant></xbrli:period></xbrli:context>
 <xbrli:context id="GW1"><xbrli:entity><xbrli:identifier scheme="urn:test">1</xbrli:identifier><xbrli:segment><xbrldi:explicitMember dimension="et-gaap:IntangibleAssetsDimension">et-gaap:IntangibleAssetsGoodwillAbstract</xbrldi:explicitMember></xbrli:segment></xbrli:entity><xbrli:period><xbrli:instant>2025-12-31</xbrli:instant></xbrli:period></xbrli:context>
 <et-gaap:Revenue contextRef="D1" unitRef="EUR">10</et-gaap:Revenue>
 <et-gaap:RevenueConsolidated contextRef="D1" unitRef="EUR">100</et-gaap:RevenueConsolidated>
 <et-gaap:RevenueConsolidated contextRef="D2" unitRef="EUR">90</et-gaap:RevenueConsolidated>
 <et-gaap:OtherIncomeConsolidated contextRef="D1" unitRef="EUR">5</et-gaap:OtherIncomeConsolidated>
 <et-gaap:RawMaterialsAndConsumablesUsedConsolidated contextRef="D1" unitRef="EUR">-50</et-gaap:RawMaterialsAndConsumablesUsedConsolidated>
 <et-gaap:TotalProfitLossConsolidated contextRef="D1" unitRef="EUR">20</et-gaap:TotalProfitLossConsolidated>
 <et-gaap:TotalProfitLossConsolidated contextRef="D2" unitRef="EUR">18</et-gaap:TotalProfitLossConsolidated>
 <et-gaap:DepreciationAndImpairmentLossReversalConsolidated contextRef="D1" unitRef="EUR">-4</et-gaap:DepreciationAndImpairmentLossReversalConsolidated>
 <et-gaap:CurrentAssetsConsolidated contextRef="I1" unitRef="EUR">70</et-gaap:CurrentAssetsConsolidated>
 <et-gaap:CurrentAssetsConsolidated contextRef="I2" unitRef="EUR">60</et-gaap:CurrentAssetsConsolidated>
 <et-gaap:NonCurrentAssetsConsolidated contextRef="I1" unitRef="EUR">30</et-gaap:NonCurrentAssetsConsolidated>
 <et-gaap:NonCurrentAssetsConsolidated contextRef="I2" unitRef="EUR">25</et-gaap:NonCurrentAssetsConsolidated>
 <et-gaap:AccountsReceivablesConsolidated contextRef="TR1" unitRef="EUR">12</et-gaap:AccountsReceivablesConsolidated>
 <et-gaap:AccountsReceivableConsolidated contextRef="TR1" unitRef="EUR">11</et-gaap:AccountsReceivableConsolidated>
 <et-gaap:AverageNumberOfEmployeesInFullTimeEquivalentUnitsConsolidated contextRef="D1">7</et-gaap:AverageNumberOfEmployeesInFullTimeEquivalentUnitsConsolidated>
 <et-gaap:IntangibleAssetsDepreciationConsolidated contextRef="GW1" unitRef="EUR">-2</et-gaap:IntangibleAssetsDepreciationConsolidated>
 <et-gaap:NetSalesByOperatingActivitiesConsolidatedTuple>
   <et-gaap:NetSalesByOperatingActivitiesConsolidatedName contextRef="D1">Jaekaubandus</et-gaap:NetSalesByOperatingActivitiesConsolidatedName>
   <et-gaap:NetSalesByOperatingActivitiesConsolidatedValue contextRef="D1" unitRef="EUR">100</et-gaap:NetSalesByOperatingActivitiesConsolidatedValue>
   <et-gaap:NetSalesByOperatingActivitiesConsolidatedValue contextRef="D2" unitRef="EUR">90</et-gaap:NetSalesByOperatingActivitiesConsolidatedValue>
 </et-gaap:NetSalesByOperatingActivitiesConsolidatedTuple>
 <et-gaap:NetSalesByGeographicalLocationInEuropeanUnionConsolidatedTuple>
   <et-gaap:NameOfCountryInEuropeanUnionConsolidated contextRef="D1">Eesti</et-gaap:NameOfCountryInEuropeanUnionConsolidated>
   <et-gaap:NetSalesInEuropeanUnionConsolidated contextRef="D1" unitRef="EUR">100</et-gaap:NetSalesInEuropeanUnionConsolidated>
   <et-gaap:NetSalesInEuropeanUnionConsolidated contextRef="D2" unitRef="EUR">90</et-gaap:NetSalesInEuropeanUnionConsolidated>
 </et-gaap:NetSalesByGeographicalLocationInEuropeanUnionConsolidatedTuple>
</xbrli:xbrl>'''
    output = BytesIO()
    with ZipFile(output, "w") as archive:
        archive.writestr("Aruanne_fixture.xbrl", xml)
    return output.getvalue()


class AdapterTests(unittest.TestCase):
    def test_scope_aware_xbrl_parses_consolidated_comparatives_and_notes(self) -> None:
        report = parse_xbrl_report(
            consolidated_xbrl_package(),
            registry_code="70000310",
            company_name="Fixture Group",
            report_year=2025,
        )
        self.assertEqual(report.accounting_basis, "consolidated")
        self.assertEqual(report.get_value(2025, "Revenue"), Decimal("100"))
        self.assertEqual(report.get_value(2024, "Revenue"), Decimal("90"))
        self.assertEqual(report.get_value(2025, "Trade debtors / receivables"), Decimal("11"))
        self.assertEqual(report.get_value(2025, "Goodwill amortisation"), Decimal("-2"))
        self.assertEqual(report.get_segments(2025, "Activity")["Retail trade"].value, Decimal("100"))
        self.assertEqual(report.get_segments(2024, "Geography")["Estonia"].value, Decimal("90"))
        self.assertTrue((report.get_source(2025, "Revenue") or "").startswith("RIK XBRL |"))

    def test_xbrl_derives_layout_two_da_from_expense_components(self) -> None:
        with ZipFile(BytesIO(consolidated_xbrl_package())) as archive:
            xml = archive.read("Aruanne_fixture.xbrl")
        xml = xml.replace(
            b'<et-gaap:DepreciationAndImpairmentLossReversalConsolidated contextRef="D1" unitRef="EUR">-4</et-gaap:DepreciationAndImpairmentLossReversalConsolidated>',
            b'<et-gaap:CostOfGoodsSoldDepreciationConsolidated contextRef="D1" unitRef="EUR">-3</et-gaap:CostOfGoodsSoldDepreciationConsolidated><et-gaap:AdministrativeExpenseDepreciationConsolidated contextRef="D1" unitRef="EUR">-1</et-gaap:AdministrativeExpenseDepreciationConsolidated>',
        )
        output = BytesIO()
        with ZipFile(output, "w") as archive:
            archive.writestr("Aruanne_fixture.xbrl", xml)
        report = parse_xbrl_report(
            output.getvalue(),
            registry_code="70000310",
            company_name="Fixture Group",
            report_year=2025,
        )
        self.assertEqual(report.get_value(2025, "D&A"), Decimal("-4"))
        self.assertIn("derived D&A", report.get_source(2025, "D&A") or "")

    def test_xbrl_replaces_unconsolidated_statement_xml_before_pdf_fallback(self) -> None:
        primary = EstonianReport(
            Path("statement.xml"),
            "RIK unconsolidated statements FY2025",
            [],
            load_terms(None),
            period_start=date(2025, 1, 1),
            period_end=date(2025, 12, 31),
            company="Fixture Group",
            accounting_basis="unconsolidated",
        )
        for item, value in {
            "Revenue": "10",
            "Reported EBIT": "2",
            "Current assets": "7",
            "Fixed assets": "3",
        }.items():
            primary.set_value(2025, item, Decimal(value), "RIK XML | fixture")
        document = CompanyDocument(
            "xbrl-2025",
            "70000310",
            "X",
            "Annual report XBRL",
            None,
            date(2026, 1, 1),
            "K",
            "A",
            2025,
            "https://example.invalid/xbrl",
            datetime(2026, 1, 1, tzinfo=timezone.utc),
        )

        class DownloadClient:
            @staticmethod
            def download_document(_document):
                return DownloadedDocument(
                    consolidated_xbrl_package(), "fixture.zip", "application/zip"
                )

        reports, warnings = add_xbrl_sources(
            DownloadClient(),
            [primary],
            [document],
            [2025],
            registry_code="70000310",
            company_name="Fixture Group",
            retry_delay_seconds=0,
        )
        self.assertEqual(reports[0].accounting_basis, "consolidated")
        self.assertEqual(reports[0].get_value(2025, "Revenue"), Decimal("100"))
        self.assertEqual(warnings, [])

    def test_xbrl_parser_rejects_web_page_downloads(self) -> None:
        with self.assertRaises(XbrlParseError):
            parse_xbrl_report(
                b"<!DOCTYPE html><html><body>rate limited</body></html>",
                registry_code="70000310",
                company_name="Fixture Group",
                report_year=2025,
            )

    def test_consolidated_statements_are_preferred(self) -> None:
        items = [
            availability("14"),
            availability("22"),
            availability("35"),
            availability("40"),
            availability("18"),
            availability("32"),
        ]
        self.assertEqual(preferred_report_types(items, 2024), ["14", "35", "18"])

    def test_income_layout_two_and_comparative_mapping(self) -> None:
        report = report_from_lines(
            "70000310",
            "Fixture Company",
            2024,
            [
                line("36", "10", "Revenue", "A1", "2000000"),
                line("36", "10", "Revenue", "A2", "1500000"),
                line("36", "110", "Operating profit", "A1", "100000"),
            ],
        )
        self.assertEqual(report.get_value(2024, "Revenue"), Decimal("2000000"))
        self.assertEqual(report.get_value(2023, "Revenue"), Decimal("1500000"))
        self.assertIn("statement 36", report.get_source(2024, "Revenue") or "")

    def test_source_report_years_include_next_filing(self) -> None:
        items = [availability("35", 2025), availability("35", 2024)]
        self.assertEqual(source_report_years(items, [2024]), [2025, 2024])

    def test_live_layout_one_rows_map_by_guarded_row_number(self) -> None:
        report = report_from_lines(
            "14847036", "Botguard OÜ", 2024, fixture_lines("14847036", 2024)
        )
        expected = {
            "Revenue": "21035457",
            "Other income": "0",
            "COGS": "-900749",
            "Reported EBIT": "-1850400",
            "D&A": "-91860",
            "Fixed assets": "0",
            "Current assets": "22235939",
            "Stocks / inventories": "0",
            "Cash and cash equivalents": "328071",
            "Debt ST": "0",
            "Debt LT": "0",
            "CAPEX": "0",
        }
        for item, value in expected.items():
            self.assertEqual(report.get_value(2024, item), Decimal(value), item)
            self.assertTrue((report.get_source(2024, item) or "").startswith("RIK XML |"))
        self.assertEqual(report.get_value(2023, "Revenue"), Decimal("8068698"))
        self.assertEqual(report.accounting_basis, "reported")

    def test_live_layout_two_and_unconsolidated_rows_map(self) -> None:
        solarstone = report_from_lines(
            "12916046", "Solarstone OÜ", 2025, fixture_lines("12916046", 2025)
        )
        self.assertEqual(solarstone.get_value(2024, "Revenue"), Decimal("1275012"))
        self.assertEqual(solarstone.get_value(2024, "COGS"), Decimal("-2040506"))
        self.assertEqual(solarstone.get_value(2024, "Reported EBIT"), Decimal("-2622236"))
        self.assertEqual(solarstone.get_value(2024, "Stocks / inventories"), Decimal("1209056"))

        blrt = report_from_lines(
            "10068499", "BLRT Grupp AS", 2025, fixture_lines("10068499", 2025)
        )
        self.assertEqual(blrt.accounting_basis, "unconsolidated")
        self.assertEqual(blrt.get_value(2024, "Fixed assets"), Decimal("245754000"))
        self.assertEqual(blrt.get_value(2024, "Current assets"), Decimal("101073000"))
        self.assertEqual(blrt.get_value(2024, "Other income"), Decimal("8538000"))
        self.assertEqual(blrt.get_value(2024, "CAPEX"), Decimal("-16892000"))

        exmet = report_from_lines(
            "11739524", "OÜ Exmet", 2025, fixture_lines("11739524", 2025)
        )
        self.assertEqual(exmet.accounting_basis, "unconsolidated")
        self.assertEqual(exmet.get_value(2025, "Revenue"), Decimal("110588963"))
        self.assertEqual(exmet.get_value(2024, "Revenue"), Decimal("101764309"))
        self.assertEqual(exmet.get_value(2025, "Debt ST"), Decimal("27316866"))
        self.assertEqual(exmet.get_value(2025, "Debt LT"), Decimal("17401296"))

    def test_consolidated_document_replaces_unconsolidated_xml_as_a_block(self) -> None:
        primary = report_from_lines(
            "10068499", "BLRT Grupp AS", 2025, fixture_lines("10068499", 2025)
        )
        fallback = EstonianReport(
            Path("blrt-2025.pdf"),
            "Consolidated annual report",
            [],
            load_terms(None),
            period_start=date(2025, 1, 1),
            period_end=date(2025, 12, 31),
            company="BLRT Grupp AS",
            accounting_basis="consolidated",
        )
        for item, value in {
            "Revenue": "440038000",
            "Reported EBIT": "36971000",
            "Fixed assets": "284692000",
            "Current assets": "304764000",
            "FTEs": "3663",
        }.items():
            fallback.set_value(2024, item, Decimal(value), "Consolidated annual report")

        replaced = replace_unconsolidated_with_consolidated(primary, fallback)
        self.assertIn(2024, replaced)
        self.assertEqual(primary.accounting_basis, "consolidated")
        self.assertEqual(primary.get_value(2024, "Revenue"), Decimal("440038000"))
        self.assertIsNone(primary.get_value(2024, "COGS"))
        self.assertEqual(primary.get_value(2024, "FTEs"), Decimal("3663"))

    def test_pdf_is_tried_before_bdoc(self) -> None:
        documents = [
            CompanyDocument(
                document_id="same",
                registry_code="70000310",
                document_type=document_type,
                document_name=document_type,
                size_bytes=size,
                status_date=date(2025, 5, 1),
                validity="K",
                report_kind="A",
                fiscal_year=2024,
                url=f"https://example.invalid/{document_type}",
                retrieved_at=datetime(2025, 5, 1, tzinfo=timezone.utc),
            )
            for document_type, size in (("D", 500), ("A", 100))
        ]
        self.assertEqual(
            [item.document_type for item in fallback_document_candidates(documents, 2024)],
            ["A", "D"],
        )

    def test_detailed_note_da_replaces_statement_level_combined_value(self) -> None:
        primary = EstonianReport(
            Path("structured.xml"),
            "RIK structured fixture",
            [],
            load_terms(None),
            company="Fixture Company",
        )
        primary.set_value(2024, "D&A", Decimal("24460"), "RIK XML | cash flow row 140")

        fallback = EstonianReport(
            Path("annual-report.pdf"),
            "Annual report fixture",
            [],
            load_terms(None),
            company="Fixture Company",
        )
        fallback.set_value(2024, "D&A", Decimal("25000"), "Cash flow statement")
        fallback.set_value(2024, "Total depreciation", Decimal("8810"), "Tangible assets note")
        fallback.set_value(2024, "Total amortisation", Decimal("19800"), "Intangible assets note")

        merge_missing(primary, fallback)

        self.assertIsNone(primary.get_value(2024, "D&A"))
        self.assertEqual(primary.get_value(2024, "Total depreciation"), Decimal("8810"))
        self.assertEqual(primary.get_value(2024, "Total amortisation"), Decimal("19800"))

    def test_partial_note_da_does_not_replace_combined_xml_value(self) -> None:
        primary = EstonianReport(
            Path("structured.xml"),
            "RIK structured fixture",
            [],
            load_terms(None),
            company="Fixture Company",
        )
        primary.set_value(2024, "D&A", Decimal("24460"), "RIK XML | cash flow row 140")
        fallback = EstonianReport(
            Path("annual-report.pdf"),
            "Annual report fixture",
            [],
            load_terms(None),
            company="Fixture Company",
        )
        fallback.set_value(2024, "Total depreciation", Decimal("8810"), "Tangible assets note")

        merge_missing(primary, fallback)

        self.assertEqual(primary.get_value(2024, "D&A"), Decimal("24460"))
        self.assertEqual(primary.get_value(2024, "Total depreciation"), Decimal("8810"))

    def test_next_annual_report_comparatives_take_priority(self) -> None:
        reports: list[EstonianReport] = []
        for report_year, values in (
            (2025, {2025: "100", 2024: "91"}),
            (2024, {2024: "80", 2023: "71"}),
            (2023, {2023: "60"}),
        ):
            report = EstonianReport(
                Path(f"AR{report_year}.xml"),
                f"AR{report_year}",
                [],
                load_terms(None),
                period_start=date(report_year, 1, 1),
                period_end=date(report_year, 12, 31),
                company="Fixture Company",
            )
            for value_year, value in values.items():
                report.set_value(
                    value_year,
                    "Revenue",
                    Decimal(value),
                    f"RIK XML | AR{report_year} "
                    f"{'current' if value_year == report_year else 'comparative'}",
                )
            reports.append(report)

        next_report = next(report for report in reports if report.year == 2025)
        current_report = next(report for report in reports if report.year == 2024)
        next_report.set_segment_value(
            2024, "Activity", "Revised label", Decimal("91"), "RIK XBRL | AR2025", "Activity", 1
        )
        current_report.set_segment_value(
            2024, "Activity", "Legacy label", Decimal("80"), "RIK XBRL | AR2024", "Activity", 1
        )

        jobs, _ = build_fill_jobs(
            reports,
            years=3,
            fill_comparative=True,
            target_years=[2025, 2024, 2023],
        )
        by_year = {job.year: job for job in jobs}
        self.assertEqual(by_year[2025].report.get_value(2025, "Revenue"), Decimal("100"))
        self.assertEqual(by_year[2024].report.get_value(2024, "Revenue"), Decimal("91"))
        self.assertEqual(by_year[2023].report.get_value(2023, "Revenue"), Decimal("71"))
        self.assertIn("AR2025 comparative", by_year[2024].report.get_source(2024, "Revenue") or "")
        self.assertIn("AR2024 comparative", by_year[2023].report.get_source(2023, "Revenue") or "")
        self.assertEqual(
            set(by_year[2024].report.get_segments(2024, "Activity")), {"Revised label"}
        )

        by_year[2024].report.set_value(2024, "FTEs", Decimal("42"), "AR2025 PDF")
        details = workbook_source_details(
            jobs,
            fill_segments=True,
            fill_cogs=True,
            fill_goodwill_amortisation=True,
        )
        detail_2024 = next(detail for detail in details if detail.startswith("FY2024"))
        self.assertIn("FY2024 ← AR2025 comparative", detail_2024)
        self.assertIn("PDF/BDOC fallback items: FTEs", detail_2024)

    def test_comparative_value_keeps_target_year_period_dates(self) -> None:
        current = EstonianReport(
            Path("AR2020.xml"),
            "AR2020",
            [],
            load_terms(None),
            period_start=date(2019, 11, 12),
            period_end=date(2020, 12, 31),
            company="Fixture Company",
        )
        current.set_value(2020, "Revenue", Decimal("80"), "AR2020 current")
        later = EstonianReport(
            Path("AR2021.xml"),
            "AR2021",
            [],
            load_terms(None),
            period_start=date(2021, 1, 1),
            period_end=date(2021, 12, 31),
            company="Fixture Company",
        )
        later.set_value(2020, "Revenue", Decimal("91"), "AR2021 comparative")

        jobs, _ = build_fill_jobs(
            [later, current],
            years=1,
            fill_comparative=True,
            target_years=[2020],
        )

        self.assertEqual(jobs[0].report.get_value(2020, "Revenue"), Decimal("91"))
        self.assertEqual(jobs[0].source_year, 2021)
        self.assertEqual(jobs[0].period_start, date(2019, 11, 12))
        self.assertEqual(jobs[0].period_end, date(2020, 12, 31))
        self.assertTrue(jobs[0].requires_annualisation)

    def test_small_geographies_are_retained_in_synthetic_rest_rows(self) -> None:
        first_records: dict[str, SegmentRecord] = {}
        last_records: dict[str, SegmentRecord] = {}
        for order in range(1, 14):
            label = f"Country {order}"
            group = "EU" if order <= 9 else "WORLD"
            first_records[label] = SegmentRecord(
                Decimal("1000000"), "Revenue note", group, order
            )
            last_records[label] = SegmentRecord(
                Decimal("2000000"), "Revenue note", group, order
            )
        for order, (label, group, value) in enumerate(
            (("Tiny EU 1", "EU", "5206"), ("Tiny EU 2", "EU", "400"), ("Tiny World", "WORLD", "1579")),
            start=14,
        ):
            first_records[label] = SegmentRecord(Decimal("0"), "Revenue note", group, order)
            last_records[label] = SegmentRecord(Decimal(value), "Revenue note", group, order)

        rows = build_segment_summary_rows(first_records, last_records)
        by_label = {row.label: row for row in rows}

        self.assertEqual(len(rows), 15)
        self.assertEqual(by_label["Rest of EU"].last_value, Decimal("5606"))
        self.assertEqual(by_label["Rest of the world"].last_value, Decimal("1579"))
        self.assertEqual(
            sum(row.last_value for row in rows),
            sum(record.value for record in last_records.values()),
        )

    def test_income_statement_da_precedes_cash_flow_adjustment(self) -> None:
        report = EstonianReport(
            Path("annual-report.pdf"),
            "Consolidated annual report",
            [
                "Konsolideeritud bilanss",
                "(eurodes)",
                "Konsolideeritud kasumiaruanne",
                "(eurodes)",
                "Põhivarade kulum ja väärtuse langus",
                "-997 340",
                "-707 976",
                "Konsolideeritud rahavoogude aruanne",
                "(eurodes)",
                "Põhivarade kulum ja väärtuse langus",
                "1 008 793",
                "707 976",
                "Konsolideeritud omakapitali muutuste aruanne",
            ],
            load_terms(None),
            period_start=date(2020, 1, 1),
            period_end=date(2020, 12, 31),
            company="Fixture Company",
            accounting_basis="consolidated",
        )

        extract_financial_values(report)

        self.assertEqual(report.get_value(2020, "D&A"), Decimal("-997340"))
        self.assertEqual(report.get_value(2019, "D&A"), Decimal("-707976"))
        self.assertIn("Income statement", report.get_source(2020, "D&A") or "")


class WorkbookTests(unittest.TestCase):
    def make_report(self, days: int) -> EstonianReport:
        report = EstonianReport(
            Path("fixture.xml"),
            "RIK structured fixture",
            [],
            load_terms(None),
            period_start=date(2024, 1, 1),
            period_end=date.fromordinal(date(2024, 1, 1).toordinal() + days - 1),
            company="Fixture Company",
        )
        for year, revenue in ((2024, "910000"), (2023, "3650000")):
            report.set_value(year, "Revenue", Decimal(revenue), f"RIK row 10 A{1 if year == 2024 else 2}")
            report.set_value(year, "Reported EBIT", Decimal("91000"), "RIK row 110")
            report.set_value(year, "D&A", Decimal("9000"), "RIK row 90")
            report.set_value(year, "Fixed assets", Decimal("1000000"), "RIK balance row 190")
            report.set_value(year, "Current assets", Decimal("500000"), "RIK balance row 100")
            report.set_value(year, "Cash and cash equivalents", Decimal("100000"), "RIK balance row 20")
        return report

    def test_annualisation_boundaries(self) -> None:
        self.assertFalse(self.make_report(90).requires_annualisation)
        self.assertTrue(self.make_report(91).requires_annualisation)
        self.assertFalse(self.make_report(365).requires_annualisation)
        self.assertFalse(self.make_report(366).requires_annualisation)

    def test_workbook_preserves_template_and_annualises_flows_only(self) -> None:
        report = self.make_report(91)
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "generated.xlsx"
            generate_workbook(
                [report],
                [2024, 2023],
                TEMPLATE,
                output,
                company_name="Fixture Company",
            )
            workbook = load_workbook(output, data_only=False)
            worksheet = workbook["Financials"]
            self.assertEqual(worksheet["P136"].value, 0.91)
            self.assertEqual(worksheet["P61"].value, "=P136/$P$133")
            self.assertEqual(worksheet["P78"].value, 1)
            self.assertEqual(worksheet["U2"].value, "CAGR 2023-2024")
            self.assertTrue(worksheet.column_dimensions["R"].hidden)
            self.assertEqual(workbook.calculation.calcMode, "auto")
            workbook.close()
            self.assertEqual(count_extended_validations(output), 2)

    def test_blank_financial_value_has_blank_confidence_level(self) -> None:
        report = self.make_report(365)
        report.set_value(2024, "Stocks / inventories", Decimal("0"), "RIK balance row 60")
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "generated.xlsx"
            generate_workbook(
                [report],
                [2024],
                TEMPLATE,
                output,
                company_name="Fixture Company",
            )
            workbook = load_workbook(output, data_only=False)
            worksheet = workbook["Financials"]
            self.assertIsNone(worksheet["P157"].value)
            self.assertIsNone(worksheet["P52"].value)
            self.assertEqual(worksheet["P42"].value, "Actual")
            workbook.close()

    def test_complete_annual_report_pdfs_are_bundled_by_year(self) -> None:
        document = CompanyDocument(
            document_id="pdf-2024",
            registry_code="70000310",
            document_type="A",
            document_name="Annual report PDF",
            size_bytes=15,
            status_date=date(2025, 5, 1),
            validity="K",
            report_kind="A",
            fiscal_year=2024,
            url="https://example.invalid/ar2024.pdf",
            retrieved_at=datetime(2025, 5, 1, tzinfo=timezone.utc),
        )

        class DownloadClient:
            def download_document(self, selected):
                self.assert_document(selected)
                return DownloadedDocument(b"%PDF-1.4\n%%EOF", "source.pdf", "application/pdf")

            @staticmethod
            def assert_document(selected):
                if selected.document_id != "pdf-2024":
                    raise AssertionError("Unexpected PDF document")

        bundle = build_annual_report_pdf_bundle(
            DownloadClient(),
            [document],
            [2024],
            company_name="Fixture Company",
        )
        self.assertEqual(bundle.included_years, (2024,))
        with ZipFile(BytesIO(bundle.content)) as archive:
            self.assertEqual(archive.namelist(), ["Fixture Company AR 2024.pdf"])
            self.assertTrue(archive.read(archive.namelist()[0]).startswith(b"%PDF-"))


if __name__ == "__main__":
    unittest.main()
