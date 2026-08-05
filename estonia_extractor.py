#!/usr/bin/env python3
"""Fill a Gain.pro-style Excel financial template from Estonian BDOC reports.

Estonian annual reports supplied here are ASiC/BDOC signed containers. The
business payload inside the Rahva Raamat examples is a generated PDF, so this
script unwraps nested BDOCs and parses the PDF text rather than expecting XML.
It keeps the same philosophy as the Polish XML extractor: write only sourceable
raw-input rows and leave the template formulas in place.
"""

from __future__ import annotations

import argparse
from copy import copy
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
import io
import json
from pathlib import Path
import re
import sys
import zipfile
from typing import Iterable

try:
    from openpyxl import load_workbook
    from openpyxl.comments import Comment
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover - user environment guard
    raise SystemExit(
        "Missing dependency: openpyxl. Install it with `pip install openpyxl`."
    ) from exc

try:
    from pypdf import PdfReader
except ImportError as exc:  # pragma: no cover - user environment guard
    raise SystemExit("Missing dependency: pypdf. Install it with `pip install pypdf`.") from exc


HEADER_ROW = 2
FINANCIALS_SHEET = "Financials"
SEGMENTS_SHEET = "Segments"
EUR_TO_EURM = Decimal("1000000")
MIN_ANNUALISATION_DAYS = 90
FULL_YEAR_DAY_COUNTS = {365, 366}
NUMERIC_LINE = re.compile(r"^[+-]?\d[\d\s]*(?:,\d+)?$")
DECIMAL_AMOUNT_IN_TEXT = re.compile(r"(?<![\d])[-+]?\d[\d\s]*(?:,\d+)")
DATE_LINE = re.compile(r"^\d{2}\.\d{2}\.\d{4}$")
YEAR_LINE = re.compile(r"^(?:19|20)\d{2}$")
DATE_RANGE_IN_TEXT = re.compile(r"(\d{2}\.\d{2}\.\d{4})\s*-\s*(\d{2}\.\d{2}\.\d{4})")
YEAR_IN_TEXT = re.compile(r"\b((?:19|20)\d{2})\b")
NOTE_HEADING = re.compile(r"^Lisa\s+\d+(?:\.(?!\d)|\s+(?!\d))", re.IGNORECASE)
NUMERIC_TOKEN = re.compile(r"^[+-]?\d+$")
NON_MONETARY_ITEMS = {"FTEs"}
GEOGRAPHY_SEGMENT = "Geography"
ACTIVITY_SEGMENT = "Activity"
SEGMENT_ROW_START = 9
SEGMENT_ROW_END = 23
SEGMENT_LABEL_COL = 2
SEGMENT_FIRST_VALUE_COL = 3
SEGMENT_LAST_VALUE_COL = 4
SEGMENT_CAGR_COL = 5
SEGMENT_BLOCKS = (
    (2, 3, 4, 5),
    (7, 8, 9, 10),
    (12, 13, 14, 15),
)
MIN_SEGMENT_DISPLAY_VALUE = Decimal("10000")
NO_LEADING_NOTE_LABELS = {
    "muud äritulud",
    "muu äritulu",
    "ärikasum",
    "kokku põhivara",
    "kokku põhivarad",
    "põhivara kokku",
    "kokku käibevara",
    "kokku käibevarad",
    "käibevara kokku",
}

GEOGRAPHY_LABELS = {
    "eesti": "Estonia",
    "soome": "Finland",
    "leedu": "Lithuania",
    "saksamaa": "Germany",
    "läti": "Latvia",
    "holland": "Netherlands",
    "madalmaad": "Netherlands",
    "küpros": "Cyprus",
    "prantsusmaa": "France",
    "rootsi": "Sweden",
    "taani": "Denmark",
    "poola": "Poland",
    "norra": "Norway",
    "türgi": "Turkey",
    "šveits": "Switzerland",
    "sveits": "Switzerland",
    "luksemburg": "Luxembourg",
    "malta": "Malta",
    "kreeka": "Greece",
    "belgia": "Belgium",
    "hispaania": "Spain",
    "iisrael": "Israel",
    "itaalia": "Italy",
    "austria": "Austria",
    "hongkong": "Hong Kong",
    "island": "Iceland",
    "marshalli saared": "Marshall Islands",
    "ukraina": "Ukraine",
    "antigua ja barbuda": "Antigua and Barbuda",
    "suurbritannia": "Great Britain",
    "araabia ühendemiraadid": "United Arab Emirates",
    "india": "India",
    "venemaa": "Russia",
    "jersey": "Jersey",
    "panama": "Panama",
    "hiina": "China",
    "ameerika ühendriigid": "US",
    "kanada": "Canada",
    "austraalia": "Australia",
    "sloveenia": "Slovenia",
    "iirimaa": "Ireland",
    "horvaatia": "Croatia",
    "singapur": "Singapore",
    "saudi araabia": "Saudi Arabia",
    "bahrein": "Bahrain",
    "serbia": "Serbia",
    "rumeenia": "Romania",
    "bulgaaria": "Bulgaria",
    "ühendkuningriigid": "Great Britain",
    "ungari": "Hungary",
}

ACTIVITY_LABELS = {
    "kauba müük": "Sales of goods",
    "kaupade müük": "Sales of goods",
    "transpordi-ja lõikamisteenus": "Transport and cutting services",
    "transpordi- ja lõikamisteenus": "Transport and cutting services",
    "üüritulu": "Rental services",
    "renditulu": "Rental services",
    "renditeenused": "Rental services",
    "jaemüük": "Retail",
    "jaekaubandus": "Retail",
    "raamatute jaemüük": "Retail",
    "hulgimüük": "Wholesale",
    "hulgikaubandus": "Wholesale",
    "toitlustus": "Catering",
    "kohvikute müük": "Catering",
    "metallide müük": "Metal sales",
    "metallkonstruktsioonide müük": "Metal structures",
    "vanametalli müük": "Scrap metal",
    "keemiatoodete müük": "Chemical products",
    "muude kaupade müük": "Other goods",
    "laevaremont": "Ship repair",
    "laevade remont": "Ship repair",
    "laevaehitus": "Shipbuilding",
    "laevade ehitus": "Shipbuilding",
    "sadamateenused": "Port services",
    "transporditeenused": "Transport services",
    "laevandus": "Shipping",
    "muud teenused": "Other services",
    "stividoriteenused": "Stevedoring services",
    "tsinkimine": "Zinc plating",
    "insenertehnilised tööd": "Engineering works",
    "insener-tehnilised tööd": "Engineering works",
    "6201 programmeerimine, tarkvaralahendused": "Programming & software",
    "62021 arvutialased konsultatsioonid": "Consulting",
    "6311 andmetöötlus, veebihosting jms": "Data processing & web hosting",
    "valuutavahetuse kasum ja teenustasud": "Currency exchange",
    "investeeringukulla müük": "Investment gold",
    "hõbeda müük": "Silver",
    "ehtekulla ja -hõbeda edasimüük sulatusse": "Scrap gold and silver",
    "maksevahendusteenus*": "Payment services",
    "maksevahendusteenus": "Payment services",
    "muud kaubad": "Other goods",
    "muud kaubad (sh plaatina ja palladium)": "Other goods",
}


DEFAULT_TERMS: dict[str, list[str]] = {
    "Revenue": ["Müügitulu", "Müügitulu kliendilepingutest"],
    "Other income": ["Muud äritulud", "Muu äritulu"],
    "COGS": [
        "Kaubad, toore, materjal ja teenused",
        "Müüdud toodangu (kaupade, teenuste) kulu",
        "Müügikulud",
        "Kokku müüdud kauba kulu",
    ],
    "Reported EBIT": ["Ärikasum (kahjum)", "Kasum äritegevusest", "Ärikasum"],
    "D&A": [
        "Põhivarade kulum ja väärtuse langus",
        "Põhivara kulum ja amortisatsioon",
        "Põhivara kulum ja väärtuse langus",
        "Põhivara amortisatsioon",
    ],
    "Fixed assets": ["Kokku põhivarad", "Kokku põhivara", "Põhivara kokku", "Põhivarad", "Põhivara"],
    "Current assets": ["Kokku käibevarad", "Kokku käibevara", "Käibevara kokku", "Käibevarad", "Käibevara"],
    "Stocks / inventories": ["Varud", "Kokku varud"],
    "Trade debtors / receivables": [
        "Nõuded ostjate vastu",
        "Nõuded ostjate vastu ja muud nõuded",
        "Nõuded klientide vastu",
        "Nõuded klientide vastu***",
    ],
    "Trade creditors / payables": ["Võlad tarnijatele", "Võlad tarnijatele ja ettemaksed"],
    "Interest-bearing debt / gross debt": [
        "Laenukohustised kokku",
        "Laenukohustised",
        "Krediitkaardikohustus",
    ],
    "Cash and cash equivalents": ["Raha", "Raha ja raha ekvivalendid", "Raha ja pangakontod"],
    "CAPEX": [
        "Tasutud materiaalsete ja immateriaalsete põhivarade soetamisel",
        "Tasutud materiaalse põhivara soetamisel",
        "Tasutud materiaalsete põhivarade soetamisel",
        "Tasutud immateriaalse põhivara soetamisel",
        "immateriaalse põhivara soetamine",
        "Põhivara soetamine",
    ],
    "FTEs": [
        "Töötajate keskmine arv taandatuna täistööajale",
        "Aruandeaasta keskmine töötajate arv (taandatuna täistööajale)",
    ],
    "Goodwill amortisation": ["Amortisatsioonikulu"],
}

@dataclass(frozen=True)
class WorkbookMapping:
    row_label: str
    item: str
    section_after: str | None
    section_before: str | None
    confidence_label: str | None = None
    scale: Decimal = EUR_TO_EURM
    blank_if_zero: bool = True
    absolute: bool = False
    optional_flag: str | None = None
    occurrence: int = 1


WORKBOOK_MAPPINGS: tuple[WorkbookMapping, ...] = (
    WorkbookMapping("Net revenue", "Revenue", "1. REPORTED FIGURES", "2. ADJUSTMENTS ", "Revenue"),
    WorkbookMapping("Other income", "Other income", "1. REPORTED FIGURES", "2. ADJUSTMENTS "),
    WorkbookMapping(
        "COGS",
        "COGS",
        "1. REPORTED FIGURES",
        "2. ADJUSTMENTS ",
        "Gross margin",
        absolute=True,
        optional_flag="fill_cogs",
    ),
    WorkbookMapping("Reported EBIT", "Reported EBIT", "1. REPORTED FIGURES", "2. ADJUSTMENTS ", "EBIT"),
    WorkbookMapping(
        "Total depreciation",
        "Total depreciation",
        "1. REPORTED FIGURES",
        "2. ADJUSTMENTS ",
        "EBITDA",
        absolute=True,
    ),
    WorkbookMapping(
        "Total amortisation",
        "Total amortisation",
        "1. REPORTED FIGURES",
        "2. ADJUSTMENTS ",
        "EBITDA",
        absolute=True,
    ),
    WorkbookMapping("D&A", "D&A", "1. REPORTED FIGURES", "2. ADJUSTMENTS ", "EBITDA", absolute=True),
    WorkbookMapping("Fixed assets", "Fixed assets", "1. REPORTED FIGURES", "2. ADJUSTMENTS ", "Total assets"),
    WorkbookMapping("Current assets", "Current assets", "1. REPORTED FIGURES", "2. ADJUSTMENTS ", "Total assets"),
    WorkbookMapping("Stocks / inventories", "Stocks / inventories", "1. REPORTED FIGURES", "2. ADJUSTMENTS ", "Inventory"),
    WorkbookMapping(
        "Trade debtors / receivables",
        "Trade debtors / receivables",
        "1. REPORTED FIGURES",
        "2. ADJUSTMENTS ",
        "Receivables",
    ),
    WorkbookMapping(
        "Trade creditors / payables",
        "Trade creditors / payables",
        "1. REPORTED FIGURES",
        "2. ADJUSTMENTS ",
        "Payables",
    ),
    WorkbookMapping(
        "Cash and cash equivalents",
        "Cash and cash equivalents",
        "1. REPORTED FIGURES",
        "2. ADJUSTMENTS ",
        "Cash & cash equivalents",
    ),
    WorkbookMapping(
        "Other financial liabilities - LT",
        "Debt LT",
        "Interest bearing debt / gross debt",
        "Title",
        "Interest-bearing debt",
        blank_if_zero=False,
    ),
    WorkbookMapping(
        "Credits and loans - ST",
        "Debt ST",
        "Interest bearing debt / gross debt",
        "Title",
        "Interest-bearing debt",
        blank_if_zero=False,
    ),
    WorkbookMapping(
        "Investments in tangible assets",
        "Investments in tangible assets",
        "1. REPORTED FIGURES",
        "2. ADJUSTMENTS ",
        "CAPEX",
        absolute=True,
    ),
    WorkbookMapping(
        "Investments in intangible assets",
        "Investments in intangible assets",
        "1. REPORTED FIGURES",
        "2. ADJUSTMENTS ",
        "CAPEX",
        absolute=True,
    ),
    WorkbookMapping("CAPEX", "CAPEX", "1. REPORTED FIGURES", "2. ADJUSTMENTS ", "CAPEX", absolute=True),
    WorkbookMapping(
        "FTEs (or employees if n/a)",
        "FTEs",
        "1. REPORTED FIGURES",
        "2. ADJUSTMENTS ",
        "FTEs",
        scale=Decimal("1"),
        blank_if_zero=False,
    ),
    WorkbookMapping(
        "Goodwill amortisation",
        "Goodwill amortisation",
        "2. ADJUSTMENTS ",
        "3. ADJUSTED FIGURES",
        absolute=True,
        optional_flag="fill_goodwill_amortisation",
    ),
    WorkbookMapping(
        "EBITDA",
        "Management EBITDA",
        "Scratchpad",
        "Interest bearing debt / gross debt",
        "Adjusted EBITDA",
        occurrence=-1,
    ),
)


@dataclass(frozen=True)
class FillJob:
    report: "EstonianReport"
    year: int
    value_year: int
    source_year: int | None = None
    period_start: date | None = None
    period_end: date | None = None

    @property
    def period_days(self) -> int | None:
        if self.period_start is None or self.period_end is None:
            return None
        return (self.period_end - self.period_start).days + 1

    @property
    def requires_annualisation(self) -> bool:
        return (
            self.period_days is not None
            and self.period_days > MIN_ANNUALISATION_DAYS
            and self.period_days not in FULL_YEAR_DAY_COUNTS
        )


@dataclass
class SegmentRecord:
    value: Decimal
    source: str
    group: str
    order: int
    is_rest: bool = False


@dataclass(frozen=True)
class SegmentSummaryRow:
    label: str
    first_value: Decimal
    last_value: Decimal
    first_source: str | None
    last_source: str | None


@dataclass(frozen=True)
class SegmentPlan:
    segmentation_of: str
    segment_by: str
    first_year: int
    last_year: int
    rows: list[SegmentSummaryRow]


@dataclass
class EstonianReport:
    source_path: Path
    payload_name: str
    lines: list[str]
    terms: dict[str, list[str]]
    money_multiplier: Decimal = Decimal("1")
    period_start: date | None = None
    period_end: date | None = None
    company: str | None = None
    accounting_basis: str = "unknown"
    values: dict[int, dict[str, Decimal]] = field(default_factory=dict)
    sources: dict[int, dict[str, str]] = field(default_factory=dict)
    segments: dict[int, dict[str, dict[str, SegmentRecord]]] = field(default_factory=dict)

    @property
    def year(self) -> int:
        if self.period_end is None:
            raise ValueError(f"Could not identify aruandeaasta lõpp in {self.source_path.name}.")
        return self.period_end.year

    @property
    def period_days(self) -> int | None:
        if self.period_start is None or self.period_end is None:
            return None
        return (self.period_end - self.period_start).days + 1

    @property
    def requires_annualisation(self) -> bool:
        return (
            self.period_days is not None
            and self.period_days > MIN_ANNUALISATION_DAYS
            and self.period_days not in FULL_YEAR_DAY_COUNTS
        )

    def set_value(self, year: int, item: str, value: Decimal | None, source: str) -> None:
        if value is None:
            return
        if item not in NON_MONETARY_ITEMS:
            value *= self.money_multiplier
        self.values.setdefault(year, {})[item] = value
        self.sources.setdefault(year, {})[item] = source

    def set_eurm_value(self, year: int, item: str, value: Decimal | None, source: str) -> None:
        if value is None:
            return
        self.values.setdefault(year, {})[item] = value * EUR_TO_EURM
        self.sources.setdefault(year, {})[item] = source

    def set_segment_value(
        self,
        year: int,
        segment_by: str,
        label: str,
        value: Decimal | None,
        source: str,
        group: str,
        order: int,
        *,
        is_rest: bool = False,
    ) -> None:
        if value is None:
            return
        records = self.segments.setdefault(year, {}).setdefault(segment_by, {})
        scaled_value = value * self.money_multiplier
        if label in records:
            records[label].value += scaled_value
            return
        records[label] = SegmentRecord(scaled_value, source, group, order, is_rest)

    def get_value(self, year: int, item: str) -> Decimal | None:
        return self.values.get(year, {}).get(item)

    def get_source(self, year: int, item: str) -> str | None:
        return self.sources.get(year, {}).get(item)

    def get_segments(self, year: int, segment_by: str) -> dict[str, SegmentRecord]:
        return self.segments.get(year, {}).get(segment_by, {})


def clean_line(value: str) -> str:
    return re.sub(r"\s+", " ", value.replace("\xa0", " ")).strip()


def normalize(value: object) -> str:
    if value is None:
        return ""
    cleaned = clean_line(str(value)).lower()
    return cleaned.replace("„", '"').replace("“", '"')


def parse_estonian_date(value: str) -> date | None:
    value = clean_line(value)
    if not DATE_LINE.fullmatch(value):
        return None
    return datetime.strptime(value, "%d.%m.%Y").date()


def infer_period_from_lines(lines: list[str], payload_name: str, source_name: str) -> tuple[date | None, date | None]:
    for line in lines[:120]:
        match = DATE_RANGE_IN_TEXT.search(clean_line(line))
        if not match:
            continue
        start = parse_estonian_date(match.group(1))
        end = parse_estonian_date(match.group(2))
        if start and end:
            return start, end

    for candidate in (payload_name, source_name, *lines[:40]):
        match = YEAR_IN_TEXT.search(clean_line(candidate))
        if match:
            year = int(match.group(1))
            return date(year, 1, 1), date(year, 12, 31)

    return None, None


def parse_amount(value: str) -> Decimal | None:
    value = clean_line(value)
    if not NUMERIC_LINE.fullmatch(value):
        return None
    try:
        return Decimal(value.replace(" ", "").replace(",", "."))
    except InvalidOperation:
        return None


def decimal_amounts_in_text(value: str) -> list[Decimal]:
    amounts: list[Decimal] = []
    for match in DECIMAL_AMOUNT_IN_TEXT.finditer(clean_line(value)):
        following = value[match.end() :].lstrip()
        if following.startswith("%"):
            continue
        parsed = parse_amount(match.group(0))
        if parsed is not None:
            amounts.append(parsed)
    return amounts


def is_amount_line(value: str) -> bool:
    return parse_amount(value) is not None


def joined_label(lines: list[str], start: int, width: int) -> str:
    return clean_line(" ".join(lines[start : start + width]))


def label_matches(value: str, term: str) -> bool:
    return normalize(value) == normalize(term)


def collect_amounts(lines: list[str], start: int, max_values: int, max_scan: int = 18) -> list[Decimal]:
    values: list[Decimal] = []
    scanned = 0
    for line in lines[start:]:
        scanned += 1
        amount = parse_amount(line)
        if amount is not None and not YEAR_LINE.fullmatch(clean_line(line)):
            values.append(amount)
            if len(values) >= max_values:
                return values
        if scanned >= max_scan:
            break
    return values


def collect_row_amounts(lines: list[str], start: int, max_scan: int = 80) -> list[Decimal]:
    values: list[Decimal] = []
    scanned = 0
    found_first_amount = False
    for line in lines[start:]:
        scanned += 1
        amount = parse_amount(line)
        if amount is not None and not YEAR_LINE.fullmatch(clean_line(line)):
            values.append(amount)
            found_first_amount = True
            continue
        if found_first_amount:
            break
        if scanned >= max_scan:
            break
    return values


def numeric_token(value: str) -> str | None:
    cleaned = value.strip().strip(",;:")
    if not NUMERIC_TOKEN.fullmatch(cleaned):
        return None
    return cleaned


def is_possible_note_token(token: str) -> bool:
    numeric = numeric_token(token)
    if numeric is None:
        return True
    if "." in token or "," in token:
        return True
    return int(numeric.lstrip("+-") or "0") <= 50


def parse_amount_tokens(tokens: list[str]) -> Decimal | None:
    if not tokens:
        return None
    numeric_tokens = [numeric_token(token) for token in tokens]
    if any(token is None for token in numeric_tokens):
        return None
    assert all(token is not None for token in numeric_tokens)
    for token in numeric_tokens[1:]:
        digits = token.lstrip("+-")
        if len(digits) != 3:
            return None
    return parse_amount(" ".join(numeric_tokens))


def candidate_amount_splits(tokens: list[str], value_count: int) -> list[tuple[list[Decimal], list[int]]]:
    candidates: list[tuple[list[Decimal], list[int]]] = []

    def recurse(position: int, groups_left: int, values: list[Decimal], lengths: list[int]) -> None:
        if groups_left == 0:
            if position == len(tokens):
                candidates.append((list(values), list(lengths)))
            return
        remaining = len(tokens) - position
        if remaining < groups_left:
            return
        max_group_size = min(3, remaining - groups_left + 1)
        for group_size in range(1, max_group_size + 1):
            value = parse_amount_tokens(tokens[position : position + group_size])
            if value is None:
                continue
            values.append(value)
            lengths.append(group_size)
            recurse(position + group_size, groups_left - 1, values, lengths)
            values.pop()
            lengths.pop()

    recurse(0, value_count, [], [])
    return candidates


def score_amount_split(candidate: tuple[list[Decimal], list[int]], dropped_tokens: int = 0) -> tuple[int, int, int, int, int, Decimal]:
    values, lengths = candidate
    maturity_table_penalty = 0
    if len(values) == 6:
        maturity_table_penalty = 0 if values[0] == values[1] + values[2] and values[3] == values[4] + values[5] else 1
    restated_shape_penalty = 0
    restated_value_penalty = 0
    if len(lengths) >= 3:
        restated_shape_penalty = 0 if lengths[-1] == lengths[-2] else 1
        restated_value_penalty = 0 if values[-1] == values[-2] else 1
    length_spread = max(lengths) - min(lengths)
    largest_abs = max(abs(value) for value in values)
    return maturity_table_penalty, restated_shape_penalty, restated_value_penalty, length_spread, dropped_tokens, largest_abs


def parse_inline_amounts_from_tail(tail: str, value_count: int, *, allow_note_prefix: bool = True) -> list[Decimal]:
    tokens = [token for token in clean_line(tail).split(" ") if token]
    if len(tokens) < value_count:
        return []
    scored_candidates: list[tuple[tuple[int, int, int, int, int, Decimal], list[Decimal]]] = []
    max_drop = min(5, len(tokens) - value_count) if allow_note_prefix else 0
    for dropped_tokens in range(max_drop + 1):
        prefix = tokens[:dropped_tokens]
        if prefix and not all(is_possible_note_token(token) for token in prefix):
            continue
        remaining = tokens[dropped_tokens:]
        for candidate in candidate_amount_splits(remaining, value_count):
            scored_candidates.append((score_amount_split(candidate, dropped_tokens), candidate[0]))
    if not scored_candidates:
        return []
    return min(scored_candidates, key=lambda item: item[0])[1]


def inline_values_for_label(
    line: str,
    terms: Iterable[str],
    max_values: int,
    table_value_count: int | None = None,
) -> tuple[list[Decimal], str] | tuple[None, None]:
    line_clean = clean_line(line)
    line_norm = normalize(line_clean)
    value_count = table_value_count or max_values
    for term in sorted(terms, key=lambda item: len(normalize(item)), reverse=True):
        term_clean = clean_line(term)
        term_norm = normalize(term_clean)
        if not term_norm:
            continue
        if line_norm != term_norm and not line_norm.startswith(term_norm + " "):
            continue
        tail = line_clean[len(term_clean) :]
        values = parse_inline_amounts_from_tail(
            tail,
            value_count,
            allow_note_prefix=term_norm not in NO_LEADING_NOTE_LABELS,
        )
        if not values and max_values == 2 and value_count == 2:
            values = parse_inline_amounts_from_tail(
                tail,
                3,
                allow_note_prefix=term_norm not in NO_LEADING_NOTE_LABELS,
            )
            if values:
                values = [values[0], values[-1]]
        if values:
            if value_count > max_values and max_values == 2:
                values = [values[0], values[-1]]
            else:
                values = values[:max_values]
            return values, term_clean
    return None, None


def inline_values_for_joined_label(
    label: str,
    terms: Iterable[str],
    max_values: int,
    table_value_count: int | None = None,
) -> tuple[list[Decimal], str] | tuple[None, None]:
    label_clean = clean_line(label)
    label_norm = normalize(label_clean)
    value_count = table_value_count or max_values
    for term in sorted(terms, key=lambda item: len(normalize(item)), reverse=True):
        term_clean = clean_line(term)
        term_norm = normalize(term_clean)
        if not term_norm:
            continue
        if label_norm != term_norm and not label_norm.startswith(term_norm + " "):
            continue
        tail = label_clean[len(term_clean) :]
        values = parse_inline_amounts_from_tail(
            tail,
            value_count,
            allow_note_prefix=term_norm not in NO_LEADING_NOTE_LABELS,
        )
        if values:
            if value_count > max_values and max_values == 2:
                values = [values[0], values[-1]]
            else:
                values = values[:max_values]
            return values, term_clean
    return None, None


def inline_table_value_count(lines: list[str], idx: int, default: int) -> int:
    context = normalize(" ".join(clean_line(line) for line in lines[max(0, idx - 35) : idx + 1]))
    if "korrigeeritud" in context:
        return max(default, 3)
    return default


def find_row_amounts_after_label(
    lines: list[str],
    terms: Iterable[str],
    *,
    max_label_lines: int = 3,
) -> tuple[list[Decimal], str] | tuple[None, None]:
    normalized_terms = list(terms)
    for idx in range(len(lines)):
        if is_amount_line(lines[idx]) or DATE_LINE.fullmatch(lines[idx]):
            continue
        inline_values, inline_label = inline_values_for_label(
            lines[idx],
            normalized_terms,
            2,
            inline_table_value_count(lines, idx, 2),
        )
        if inline_values and inline_label:
            return inline_values, inline_label
        for width in range(1, max_label_lines + 1):
            if idx + width > len(lines):
                continue
            label = joined_label(lines, idx, width)
            if any(label_matches(label, term) for term in normalized_terms):
                values = collect_row_amounts(lines, idx + width)
                if values:
                    return values, label
    return None, None


def find_values_after_label(
    lines: list[str],
    terms: Iterable[str],
    *,
    max_values: int = 2,
    max_label_lines: int = 3,
) -> tuple[list[Decimal], str] | tuple[None, None]:
    normalized_terms = list(terms)
    for idx in range(len(lines)):
        if is_amount_line(lines[idx]) or DATE_LINE.fullmatch(lines[idx]):
            continue
        inline_values, inline_label = inline_values_for_label(
            lines[idx],
            normalized_terms,
            max_values,
            inline_table_value_count(lines, idx, max_values),
        )
        if inline_values and inline_label:
            return inline_values, inline_label
        for width in range(1, max_label_lines + 1):
            if idx + width > len(lines):
                continue
            label = joined_label(lines, idx, width)
            inline_values, inline_label = inline_values_for_joined_label(
                label,
                normalized_terms,
                max_values,
                inline_table_value_count(lines, idx, max_values),
            )
            if inline_values and inline_label:
                return inline_values, inline_label
            if any(label_matches(label, term) for term in normalized_terms):
                values = collect_amounts(lines, idx + width, max_values=max_values)
                if values:
                    return values, label
    return None, None


def find_all_values_after_label(
    lines: list[str],
    terms: Iterable[str],
    *,
    max_values: int = 2,
    max_label_lines: int = 3,
) -> list[tuple[list[Decimal], str]]:
    matches: list[tuple[list[Decimal], str]] = []
    for idx in range(len(lines)):
        if is_amount_line(lines[idx]) or DATE_LINE.fullmatch(lines[idx]):
            continue
        inline_values, inline_label = inline_values_for_label(
            lines[idx],
            terms,
            max_values,
            inline_table_value_count(lines, idx, max_values),
        )
        if inline_values and inline_label:
            matches.append((inline_values, inline_label))
            continue
        for width in range(1, max_label_lines + 1):
            if idx + width > len(lines):
                continue
            label = joined_label(lines, idx, width)
            if any(label_matches(label, term) for term in terms):
                values = collect_amounts(lines, idx + width, max_values=max_values)
                if values:
                    matches.append((values, label))
                break
    return matches


def date_block(lines: list[str], year: int) -> list[str] | None:
    marker = f"31.12.{year}"
    starts = [idx for idx, line in enumerate(lines) if clean_line(line) == marker]
    if not starts:
        return None
    start = starts[0]
    end = len(lines)
    for idx in range(start + 1, len(lines)):
        if DATE_LINE.fullmatch(clean_line(lines[idx])):
            end = idx
            break
    return lines[start:end]


def section_between(lines: list[str], start: int | None, end: int | None) -> list[str]:
    if start is None:
        return []
    return lines[start : end or len(lines)]


def find_main_heading(lines: list[str], headings: Iterable[str], after: int = 0) -> int | None:
    normalized = {normalize(heading) for heading in headings}
    for idx in range(after, len(lines)):
        if normalize(lines[idx]) not in normalized:
            continue
        lookahead = lines[idx : idx + 6]
        if any(is_money_unit_marker(line) for line in lookahead):
            return idx
    return None


def is_money_unit_marker(value: str) -> bool:
    normalized = normalize(value)
    return normalized.startswith("(") and "euro" in normalized


def detect_money_multiplier(lines: list[str]) -> Decimal:
    return Decimal("1000") if any("tuhandetes eurodes" in normalize(line) for line in lines) else Decimal("1")


def find_first_after(lines: list[str], terms: Iterable[str], after: int = 0) -> int | None:
    normalized = {normalize(term) for term in terms}
    for idx in range(after, len(lines)):
        if normalize(lines[idx]) in normalized:
            return idx
    return None


def find_first_prefix_after(lines: list[str], terms: Iterable[str], after: int = 0) -> int | None:
    normalized = [normalize(term) for term in terms]
    for idx in range(after, len(lines)):
        line = normalize(lines[idx])
        if any(line == term or line.startswith(term + " ") for term in normalized):
            return idx
    return None


def find_first_containing(lines: list[str], fragments: Iterable[str], after: int = 0) -> int | None:
    normalized_fragments = [normalize(fragment) for fragment in fragments]
    for idx in range(after, len(lines)):
        line = normalize(lines[idx])
        if any(fragment in line for fragment in normalized_fragments):
            return idx
    return None


def find_note_section(lines: list[str], title_contains: str, after: int) -> list[str]:
    title_norm = normalize(title_contains)
    start: int | None = None
    for idx in range(after, len(lines)):
        line_norm = normalize(lines[idx])
        if NOTE_HEADING.match(lines[idx]) and title_norm in line_norm:
            start = idx
            break
    if start is None:
        return []

    end = len(lines)
    for idx in range(start + 1, len(lines)):
        if NOTE_HEADING.match(lines[idx]):
            end = idx
            break
    return lines[start:end]


def find_first_note_section(lines: list[str], title_options: Iterable[str], after: int) -> list[str]:
    for title in title_options:
        section = find_note_section(lines, title, after)
        if section:
            return section
    return []


def extract_pdf_from_container(path: Path) -> tuple[bytes, str]:
    data = path.read_bytes()
    if path.suffix.lower() == ".pdf":
        return data, path.name

    if not zipfile.is_zipfile(io.BytesIO(data)):
        raise ValueError(f"{path.name} is neither a PDF nor a ZIP/BDOC container.")

    return extract_pdf_from_zip(data, path.name)


def extract_pdf_from_zip(data: bytes, container_name: str) -> tuple[bytes, str]:
    with zipfile.ZipFile(io.BytesIO(data)) as zf:
        nested_bdocs = [
            name
            for name in zf.namelist()
            if name.lower().endswith((".bdoc", ".ddoc", ".asice")) and not name.startswith("META-INF/")
        ]
        for name in nested_bdocs:
            try:
                return extract_pdf_from_zip(zf.read(name), name)
            except Exception:
                continue

        pdfs = [
            name
            for name in zf.namelist()
            if name.lower().endswith(".pdf") and not name.startswith("META-INF/")
        ]
        if not pdfs:
            raise ValueError(f"No annual-report PDF found inside {container_name}.")

        preferred = [
            name
            for name in pdfs
            if "aruanne" in name.lower() and "audiitor" not in name.lower() and "otsus" not in name.lower()
        ]
        selected = preferred[0] if preferred else pdfs[0]
        return zf.read(selected), selected


def pdf_text_lines(pdf_bytes: bytes) -> list[str]:
    reader = PdfReader(io.BytesIO(pdf_bytes))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    return [line for line in (raw.replace("\xa0", " ").strip() for raw in text.splitlines()) if line]


def detect_document_accounting_basis(lines: list[str]) -> str:
    """Identify a consolidated report without assuming every filing is a group filing."""
    normalized_lines = [normalize(line) for line in lines]
    if any(
        "konsolideeritud raamatupidamise aastaaruanne" in line
        or "consolidated financial statements" in line
        for line in normalized_lines
    ):
        return "consolidated"
    if any("konsolideerimata" in line or "unconsolidated" in line for line in normalized_lines):
        return "unconsolidated"
    return "reported"


def load_terms(mapping_json: Path | None) -> dict[str, list[str]]:
    terms = {item: list(values) for item, values in DEFAULT_TERMS.items()}
    if mapping_json is None:
        return terms

    with mapping_json.open("r", encoding="utf-8") as handle:
        data = json.load(handle)
    for item in data.get("items", []):
        canonical = item.get("canonical_item")
        estonian_terms = item.get("estonian_terms") or []
        if not canonical:
            continue
        terms.setdefault(canonical, [])
        for term in estonian_terms:
            if term and all(normalize(term) != normalize(existing) for existing in terms[canonical]):
                terms[canonical].append(term)
    return terms


def parse_report_from_pdf_bytes(
    source_path: Path,
    payload_name: str,
    pdf_bytes: bytes,
    terms: dict[str, list[str]],
) -> EstonianReport:
    lines = pdf_text_lines(pdf_bytes)
    report = EstonianReport(
        source_path,
        payload_name,
        lines,
        terms,
        money_multiplier=detect_money_multiplier(lines),
        accounting_basis=detect_document_accounting_basis(lines),
    )
    report.period_start = find_following_date(lines, "aruandeaasta algus:")
    report.period_end = find_following_date(lines, "aruandeaasta lõpp:")
    if report.period_start is None or report.period_end is None:
        inferred_start, inferred_end = infer_period_from_lines(lines, payload_name, source_path.name)
        report.period_start = report.period_start or inferred_start
        report.period_end = report.period_end or inferred_end
    report.company = find_following_text(lines, "ärinimi:")
    extract_financial_values(report)
    return report


def parse_report(path: Path, terms: dict[str, list[str]]) -> EstonianReport:
    pdf_bytes, payload_name = extract_pdf_from_container(path)
    return parse_report_from_pdf_bytes(path, payload_name, pdf_bytes, terms)


def parse_reports_from_input(path: Path, terms: dict[str, list[str]]) -> list[EstonianReport]:
    if path.suffix.lower() != ".zip":
        return [parse_report(path, terms)]

    reports: list[EstonianReport] = []
    with zipfile.ZipFile(path) as zf:
        pdf_names = [
            name
            for name in zf.namelist()
            if name.lower().endswith(".pdf")
            and not name.startswith("__MACOSX/")
            and "audiitor" not in name.lower()
            and "otsus" not in name.lower()
        ]
        nested_containers = [
            name
            for name in zf.namelist()
            if name.lower().endswith((".bdoc", ".ddoc", ".asice"))
            and not name.startswith("__MACOSX/")
        ]

        for name in sorted(pdf_names):
            reports.append(
                parse_report_from_pdf_bytes(
                    Path(f"{path.name}/{name}"),
                    name,
                    zf.read(name),
                    terms,
                )
            )

        for name in sorted(nested_containers):
            pdf_bytes, payload_name = extract_pdf_from_zip(zf.read(name), name)
            reports.append(
                parse_report_from_pdf_bytes(
                    Path(f"{path.name}/{name}"),
                    payload_name,
                    pdf_bytes,
                    terms,
                )
            )

    if not reports:
        raise ValueError(f"No annual-report PDFs or BDOC/DDOC containers found inside {path.name}.")
    return reports


def find_following_date(lines: list[str], label: str) -> date | None:
    target = normalize(label)
    for idx, line in enumerate(lines):
        if normalize(line) != target:
            continue
        for candidate in lines[idx + 1 : idx + 5]:
            parsed = parse_estonian_date(candidate)
            if parsed:
                return parsed
    return None


def find_following_text(lines: list[str], label: str) -> str | None:
    target = normalize(label)
    for idx, line in enumerate(lines):
        if normalize(line) == target:
            for candidate in lines[idx + 1 : idx + 5]:
                if candidate:
                    return candidate
    return None


def extract_financial_values(report: EstonianReport) -> None:
    lines = report.lines
    balance_start = find_main_heading(
        lines,
        (
            "Konsolideeritud bilanss",
            "Bilanss",
            "Konsolideeritud finantsseisundi aruanne",
            "Finantsseisundi aruanne",
        ),
        after=0,
    )
    income_start = find_main_heading(
        lines,
        ("Konsolideeritud kasumiaruanne", "Kasumiaruanne"),
        after=balance_start or 0,
    )
    cashflow_start = find_main_heading(
        lines,
        ("Konsolideeritud rahavoogude aruanne", "Rahavoogude aruanne"),
        after=income_start or 0,
    )
    statement_end_anchor = cashflow_start or income_start or balance_start or 100
    notes_start = find_first_after(lines, ("Lisa 1 Arvestuspõhimõtted",), after=statement_end_anchor)
    equity_start = find_first_after(
        lines,
        ("Konsolideeritud omakapitali muutuste aruanne", "Omakapitali muutuste aruanne"),
        after=statement_end_anchor,
    )
    cashflow_end_candidates = [idx for idx in (notes_start, equity_start) if idx is not None]
    cashflow_end = min(cashflow_end_candidates) if cashflow_end_candidates else None

    balance = section_between(lines, balance_start, income_start)
    income = section_between(lines, income_start, cashflow_start)
    cashflow = section_between(lines, cashflow_start, cashflow_end)
    notes_after = notes_start or cashflow_end or 0

    current_year = report.year
    prior_year = current_year - 1

    assign_management_ebitda(report, lines)
    assign_dual(report, income, "Revenue", "Income statement")
    assign_dual(report, income, "Other income", "Income statement")
    assign_dual(report, income, "COGS", "Income statement")
    assign_dual(report, income, "Reported EBIT", "Income statement")
    # EBITDA must add back the D&A expense included in EBIT. The similarly named
    # cash-flow adjustment can differ because it may include other non-cash
    # movements, so use it only when the income statement has no D&A line.
    assign_dual(report, income, "D&A", "Income statement")
    assign_note_dual_if_missing(report, cashflow, "D&A", "Cash flow statement")
    assign_dual(
        report,
        balance,
        "Fixed assets",
        "Balance sheet",
        terms_override=("Kokku põhivarad", "Kokku põhivara", "Põhivara kokku"),
    )
    assign_dual(
        report,
        balance,
        "Current assets",
        "Balance sheet",
        terms_override=("Kokku käibevarad", "Kokku käibevara", "Käibevara kokku"),
    )
    assign_dual(report, balance, "Stocks / inventories", "Balance sheet")
    assign_dual(report, balance, "Cash and cash equivalents", "Balance sheet")
    assign_balance_debt_split(report, balance)
    assign_capex(report, cashflow)
    assign_balance_debt_fallback(report, balance)

    receivables_note = find_first_note_section(
        lines,
        (
            "Nõuded ja ettemaksed",
            "Nõuded ja tehtud ettemaksed",
            "Nõuded ostjate vastu ja muud nõuded",
        ),
        notes_after,
    )
    payables_note = find_first_note_section(
        lines,
        (
            "Võlad ja ettemaksed",
            "Võlad ja saadud ettemaksed",
            "Võlad tarnijatele ja ettemaksed",
        ),
        notes_after,
    )
    debt_note = find_note_section(lines, "Laenukohustised", notes_after)
    employee_note = find_note_section(lines, "Tööjõukulud", notes_after)
    tangible_note = find_note_section(lines, "Materiaalsed põhivarad", notes_after)
    intangible_note = find_note_section(lines, "Immateriaalsed põhivarad", notes_after)
    revenue_note = find_note_section(lines, "Müügitulu", notes_after)

    assign_revenue_geography_segments(report, revenue_note)
    assign_revenue_activity_segments(report, revenue_note)
    assign_by_date_block(report, receivables_note, "Trade debtors / receivables", "Receivables note", (current_year, prior_year))
    assign_by_date_block(report, payables_note, "Trade creditors / payables", "Payables note", (current_year, prior_year))
    assign_by_date_block(report, debt_note, "Interest-bearing debt / gross debt", "Debt note", (current_year, prior_year))
    assign_receivables_maturity_note(report, receivables_note)
    assign_note_dual_if_missing(report, payables_note, "Trade creditors / payables", "Payables note")
    assign_asset_note_movements(
        report,
        tangible_note,
        "Total depreciation",
        "Investments in tangible assets",
        "Tangible assets note",
    )
    assign_asset_note_movements(
        report,
        intangible_note,
        "Total amortisation",
        "Investments in intangible assets",
        "Intangible assets note",
    )
    assign_dual(report, employee_note, "FTEs", "Employee note")
    assign_goodwill_amortisation(report, intangible_note)


def assign_dual(
    report: EstonianReport,
    section: list[str],
    item: str,
    source_prefix: str,
    terms_override: Iterable[str] | None = None,
) -> None:
    if not section:
        return
    search_terms = list(terms_override or report.terms.get(item, []))
    values, label = find_values_after_label(section, search_terms, max_values=2)
    if not values or label is None:
        return
    years = (report.year, report.year - 1)
    for year, value in zip(years, values):
        report.set_value(year, item, value, f"{source_prefix}: {label}")


def assign_management_ebitda(report: EstonianReport, lines: list[str]) -> None:
    for idx, line in enumerate(lines):
        if not normalize(line).startswith("ebitda ="):
            continue
        values: list[Decimal] = []
        for candidate in lines[idx : idx + 8]:
            candidate_values = decimal_amounts_in_text(candidate)
            if len(candidate_values) >= 2:
                values = candidate_values[:2]
                break
        if not values:
            return
        for year, value in zip((report.year, report.year - 1), values):
            report.set_eurm_value(year, "Management EBITDA", value, "Management report KPI table: EBITDA")
        return


def geography_group_heading(value: str) -> str | None:
    normalized = normalize(value)
    if normalized in {"müük euroopa liidu riikidele", "müügid euroopa liidus:"}:
        return "EU"
    if normalized in {
        "müük väljapoole euroopa liidu riike",
        "müük väljaspool euroopa liidu riike",
        "müügid mujal:",
    }:
        return "WORLD"
    return None


def geography_group_total(value: str) -> str | None:
    normalized = normalize(value)
    eu_totals = {
        "müük euroopa liidu riikidele, kokku",
        "müük euroopa liidu riikidele kokku",
        "kokku müügid euroopa liidus",
    }
    if normalized in eu_totals or any(normalized.startswith(total + " ") for total in eu_totals):
        return "EU"
    world_totals = {
        "müük väljapoole euroopa liidu riike, kokku",
        "müük väljapoole euroopa liidu riike kokku",
        "müük väljaspool euroopa liidu riike, kokku",
        "müük väljaspool euroopa liidu riike kokku",
        "kokku müügid mujal",
    }
    if normalized in world_totals or any(normalized.startswith(total + " ") for total in world_totals):
        return "WORLD"
    return None


def geography_explicit_rest_group(value: str) -> str | None:
    normalized = normalize(value)
    if "euroopa liidu riikidele, muud" in normalized:
        return "EU"
    if "väljapoole euroopa liidu riike, muud" in normalized or "väljaspool euroopa liidu riike, muud" in normalized:
        return "WORLD"
    return None


def geography_label(value: str, current_group: str) -> tuple[str, str, bool] | tuple[None, None, None]:
    normalized = normalize(value)
    if "kokku" in normalized:
        return None, None, None
    if "euroopa liidu riikidele, muud" in normalized:
        return "Rest of EU", "EU", True
    if "väljapoole euroopa liidu riike, muud" in normalized or "väljaspool euroopa liidu riike, muud" in normalized:
        return "Rest of the world", "WORLD", True
    if normalized.startswith("müük ") or normalized.startswith("müügid "):
        return None, None, None
    return GEOGRAPHY_LABELS.get(normalized, value), current_group, False


def rest_label_for_group(group: str) -> str:
    if group == "EU":
        return "Rest of EU"
    if group == "WORLD":
        return "Rest of the world"
    return "Other"


def inline_values_after_prefix(line: str, prefix: str, value_count: int = 2) -> list[Decimal]:
    line_clean = clean_line(line)
    prefix_clean = clean_line(prefix)
    line_norm = normalize(line_clean)
    prefix_norm = normalize(prefix_clean)
    if line_norm != prefix_norm and not line_norm.startswith(prefix_norm + " "):
        return []
    return parse_inline_amounts_from_tail(
        line_clean[len(prefix_clean) :],
        value_count,
        allow_note_prefix=False,
    )


def inline_segment_row(
    line: str,
    labels: dict[str, str],
    value_count: int = 2,
) -> tuple[str, list[Decimal], str] | tuple[None, None, None]:
    for source_label, display_label in sorted(labels.items(), key=lambda item: len(normalize(item[0])), reverse=True):
        values = inline_values_after_prefix(line, source_label, value_count)
        if values:
            return display_label, values, source_label
    return None, None, None


def assign_revenue_geography_segments(report: EstonianReport, revenue_note: list[str]) -> None:
    if not revenue_note:
        return

    start = find_first_after(revenue_note, ("Müügitulu geograafiliste piirkondade lõikes",))
    if start is None:
        start = find_first_containing(revenue_note, ("müügitulu jagunemine riigiti",))
    if start is None:
        return
    end = len(revenue_note)
    for idx in range(start + 1, len(revenue_note)):
        line_norm = normalize(revenue_note[idx])
        if (
            line_norm == "müügitulu tegevusalade lõikes"
            or line_norm.startswith("lisa ")
            or NOTE_HEADING.match(revenue_note[idx])
        ):
            end = idx
            break

    explicit_rest_groups: set[str] = set()
    explicit_country_count = 0
    scan_group = ""
    for idx in range(start + 1, end):
        line = revenue_note[idx]
        group = geography_group_heading(line)
        if group is not None:
            scan_group = group
            continue
        rest_group = geography_explicit_rest_group(line)
        if rest_group is not None:
            explicit_rest_groups.add(rest_group)
            continue
        if geography_group_total(line) is not None:
            continue
        if is_amount_line(line) or YEAR_LINE.fullmatch(clean_line(line)) or normalize(line) == "lisa nr":
            continue
        inline_label, _inline_values, _source_label = inline_segment_row(line, GEOGRAPHY_LABELS)
        if inline_label is not None:
            label, label_group = inline_label, scan_group
        else:
            label, label_group, _ = geography_label(line, scan_group)
        if label is not None and label_group is not None:
            explicit_country_count += 1

    group_totals: dict[str, tuple[list[Decimal], str]] = {}
    group_total_prefixes = {
        "EU": (
            "Müük Euroopa Liidu riikidele, kokku",
            "Müük Euroopa Liidu riikidele kokku",
            "Kokku müügid Euroopa Liidus",
        ),
        "WORLD": (
            "Müük väljapoole Euroopa Liidu riike, kokku",
            "Müük väljapoole Euroopa Liidu riike kokku",
            "Müük väljaspool Euroopa Liidu riike, kokku",
            "Müük väljaspool Euroopa Liidu riike kokku",
            "Kokku müügid mujal",
        ),
    }
    for idx in range(start + 1, end):
        group_total = geography_group_total(revenue_note[idx])
        if group_total is None or group_total in explicit_rest_groups:
            continue
        values: list[Decimal] = []
        for prefix in group_total_prefixes[group_total]:
            values = inline_values_after_prefix(revenue_note[idx], prefix)
            if values:
                break
        if not values:
            values = collect_amounts(revenue_note, idx + 1, max_values=2, max_scan=6)
        if values:
            group_totals[group_total] = (values, revenue_note[idx])

    current_group = ""
    order = 0
    years = (report.year, report.year - 1)
    for idx in range(start + 1, end):
        line = revenue_note[idx]
        if is_amount_line(line) or YEAR_LINE.fullmatch(clean_line(line)) or normalize(line) == "lisa nr":
            continue
        if geography_group_total(line) is not None:
            continue

        group = geography_group_heading(line)
        if group is not None:
            current_group = group
            continue

        inline_label, inline_values, source_label = inline_segment_row(line, GEOGRAPHY_LABELS)
        if inline_label is not None and inline_values is not None and source_label is not None:
            label, label_group, is_rest = inline_label, current_group, False
            values = inline_values
        else:
            label, label_group, is_rest = geography_label(line, current_group)
            values = collect_amounts(revenue_note, idx + 1, max_values=2, max_scan=6)
        if label is None or label_group is None:
            continue
        if not values:
            continue

        order += 1
        for year, value in zip(years, values):
            report.set_segment_value(
                year,
                GEOGRAPHY_SEGMENT,
                label,
                value,
                f"Revenue note geography: {line}",
                label_group,
                order,
                is_rest=is_rest,
            )

    for group, (values, line) in group_totals.items():
        rest_label = rest_label_for_group(group)
        order += 1
        for idx, year in enumerate(years):
            if idx >= len(values):
                continue
            records = report.get_segments(year, GEOGRAPHY_SEGMENT)
            selected_value = sum(
                record.value
                for record in records.values()
                if record.group == group and not record.is_rest
            )
            rest_value = (values[idx] * report.money_multiplier) - selected_value
            if rest_value == 0:
                continue
            report.set_segment_value(
                year,
                GEOGRAPHY_SEGMENT,
                rest_label,
                rest_value / report.money_multiplier,
                f"Revenue note geography: {line}",
                group,
                order,
                is_rest=True,
            )


def activity_label(value: str) -> str | None:
    normalized = normalize(value)
    if "kokku" in normalized or normalized == "lisa nr":
        return None
    if normalized.startswith("müügitulu "):
        return None
    return ACTIVITY_LABELS.get(normalized, value)


def assign_revenue_activity_segments(report: EstonianReport, revenue_note: list[str]) -> None:
    if not revenue_note:
        return

    start = find_first_after(revenue_note, ("Müügitulu tegevusalade lõikes",))
    if start is None:
        start = find_first_containing(revenue_note, ("müügitulu jagunemine tegevusalade lõikes",))
    if start is None:
        return
    end = len(revenue_note)
    for idx in range(start + 1, len(revenue_note)):
        line_norm = normalize(revenue_note[idx])
        if (
            line_norm.startswith("kokku müügitulu")
            or "müügitulu jagunemine riigiti" in line_norm
            or NOTE_HEADING.match(revenue_note[idx])
        ):
            end = idx
            break

    order = 0
    years = (report.year, report.year - 1)
    for idx in range(start + 1, end):
        line = revenue_note[idx]
        if is_amount_line(line) or YEAR_LINE.fullmatch(clean_line(line)):
            continue
        inline_label, inline_values, _source_label = inline_segment_row(line, ACTIVITY_LABELS)
        if inline_label is not None and inline_values is not None:
            label = inline_label
            values = inline_values
        else:
            label = activity_label(line)
            if label is None:
                continue
            values = collect_amounts(revenue_note, idx + 1, max_values=2, max_scan=6)
        if not values:
            continue

        order += 1
        for year, value in zip(years, values):
            report.set_segment_value(
                year,
                ACTIVITY_SEGMENT,
                label,
                value,
                f"Revenue note activity: {line}",
                ACTIVITY_SEGMENT,
                order,
            )


def assign_by_date_block(
    report: EstonianReport,
    section: list[str],
    item: str,
    source_prefix: str,
    years: Iterable[int],
) -> None:
    if not section:
        return
    for year in years:
        block = date_block(section, year)
        if not block:
            continue
        values, label = find_values_after_label(block, report.terms.get(item, []), max_values=1)
        if values and label is not None:
            report.set_value(year, item, values[0], f"{source_prefix}: {label}")


def assign_note_dual_if_missing(
    report: EstonianReport,
    section: list[str],
    item: str,
    source_prefix: str,
    terms_override: Iterable[str] | None = None,
) -> None:
    if not section:
        return
    values, label = find_values_after_label(section, terms_override or report.terms.get(item, []), max_values=2)
    if not values or label is None:
        return
    for year, value in zip((report.year, report.year - 1), values):
        if report.get_value(year, item) is None:
            report.set_value(year, item, value, f"{source_prefix}: {label}")


def assign_receivables_maturity_note(report: EstonianReport, section: list[str]) -> None:
    if not section:
        return
    terms = ("Nõuded klientide vastu***", "Nõuded klientide vastu", "Nõuded ostjate vastu")
    for line in section:
        values, label = inline_values_for_label(line, terms, max_values=6, table_value_count=6)
        if not values or label is None or len(values) < 4:
            continue
        for year, value in ((report.year, values[0]), (report.year - 1, values[3])):
            if report.get_value(year, "Trade debtors / receivables") is None:
                report.set_value(
                    year,
                    "Trade debtors / receivables",
                    value,
                    f"Receivables note: {label}",
                )
        return


def balance_subsection(lines: list[str], start_label: str, end_label: str) -> list[str]:
    start = find_first_after(lines, (start_label,))
    if start is None:
        return []
    end = find_first_prefix_after(lines, (end_label,), after=start + 1)
    return lines[start : end or len(lines)]


def assign_balance_debt_split(report: EstonianReport, balance: list[str]) -> None:
    if not balance:
        return

    sections = (
        ("Debt ST", "Lühiajalised kohustised", "Kokku lühiajalised kohustised"),
        ("Debt LT", "Pikaajalised kohustised", "Kokku pikaajalised kohustised"),
    )
    for item, start_label, end_label in sections:
        section = balance_subsection(balance, start_label, end_label)
        if not section:
            continue
        values, label = find_values_after_label(section, ("Laenukohustised",), max_values=2)
        if not values or label is None:
            continue
        for year, value in zip((report.year, report.year - 1), values):
            report.set_value(year, item, value, f"Balance sheet {start_label}: {label}")


def assign_asset_note_movements(
    report: EstonianReport,
    note: list[str],
    expense_item: str,
    purchase_item: str,
    source_prefix: str,
) -> None:
    if not note:
        return

    for fiscal_year in (report.year, report.year - 1):
        opening = f"31.12.{fiscal_year - 1}"
        closing = f"31.12.{fiscal_year}"
        opening_indices = [idx for idx, line in enumerate(note) if clean_line(line) == opening]
        if not opening_indices:
            continue
        start = opening_indices[-1]
        end = len(note)
        for idx in range(start + 1, len(note)):
            if clean_line(note[idx]) == closing:
                end = idx
                break
        block = note[start:end]

        expense_values, expense_label = find_row_amounts_after_label(block, ("Amortisatsioonikulu",))
        if expense_values and expense_label is not None:
            report.set_value(
                fiscal_year,
                expense_item,
                expense_values[-1],
                f"{source_prefix}: {expense_label}",
            )

        purchase_values, purchase_label = find_row_amounts_after_label(block, ("Ostud ja parendused",))
        if purchase_values and purchase_label is not None:
            report.set_value(
                fiscal_year,
                purchase_item,
                purchase_values[-1],
                f"{source_prefix}: {purchase_label}",
            )


def assign_balance_debt_fallback(report: EstonianReport, balance: list[str]) -> None:
    if not balance:
        return
    matches = find_all_values_after_label(balance, ("Laenukohustised",), max_values=2)
    if not matches:
        return
    totals = [Decimal("0"), Decimal("0")]
    found = [False, False]
    for values, _label in matches:
        for idx, value in enumerate(values[:2]):
            totals[idx] += value
            found[idx] = True

    for idx, year in enumerate((report.year, report.year - 1)):
        if found[idx] and report.get_value(year, "Interest-bearing debt / gross debt") is None:
            report.set_value(
                year,
                "Interest-bearing debt / gross debt",
                totals[idx],
                "Balance sheet: Laenukohustised",
            )


def assign_capex(report: EstonianReport, cashflow: list[str]) -> None:
    if not cashflow:
        return
    combined_terms = [
        "Tasutud materiaalsete ja immateriaalsete põhivarade soetamisel",
        "Tasutud materiaalsete põhivarade soetamisel",
        "Tasutud materiaalse põhivara soetamisel",
        "Põhivara soetamine",
    ]
    values, label = find_values_after_label(cashflow, combined_terms, max_values=2)
    if values and label:
        for year, value in zip((report.year, report.year - 1), values):
            report.set_value(year, "CAPEX", value, f"Cash flow statement: {label}")
        return

    split_terms = [
        "Tasutud materiaalse põhivara soetamisel",
        "Tasutud immateriaalse põhivara soetamisel",
        "immateriaalse põhivara soetamine",
    ]
    matches = find_all_values_after_label(cashflow, split_terms, max_values=2)
    if not matches:
        return
    totals = [Decimal("0"), Decimal("0")]
    found = [False, False]
    labels: list[str] = []
    for values_for_row, label_for_row in matches:
        labels.append(label_for_row)
        for idx, value in enumerate(values_for_row[:2]):
            totals[idx] += value
            found[idx] = True
    for idx, year in enumerate((report.year, report.year - 1)):
        if found[idx]:
            report.set_value(year, "CAPEX", totals[idx], "Cash flow statement: " + "; ".join(labels))


def assign_goodwill_amortisation(report: EstonianReport, note: list[str]) -> None:
    if not note or not any("firmaväärtus" in normalize(line) for line in note[:30]):
        return

    for fiscal_year in (report.year, report.year - 1):
        opening = f"31.12.{fiscal_year - 1}"
        closing = f"31.12.{fiscal_year}"
        opening_indices = [idx for idx, line in enumerate(note) if clean_line(line) == opening]
        if not opening_indices:
            continue
        start = opening_indices[-1]
        end = len(note)
        for idx in range(start + 1, len(note)):
            if clean_line(note[idx]) == closing:
                end = idx
                break
        block = note[start:end]
        values, label = find_values_after_label(block, ("Amortisatsioonikulu",), max_values=1)
        if values and label is not None:
            report.set_value(
                fiscal_year,
                "Goodwill amortisation",
                values[0],
                f"Intangible assets note: {label} / Firmaväärtus",
            )


def find_year_column(ws, year: int) -> int:
    for cell in ws[HEADER_ROW]:
        if cell.value == year or str(cell.value).strip() == str(year):
            return cell.column
    raise ValueError(f"Could not find year {year} in row {HEADER_ROW} of the Financials sheet.")


def header_year(value: object) -> int | None:
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, str) and re.fullmatch(r"\d{4}", value.strip()):
        return int(value.strip())
    return None


def row_label(ws, row: int) -> str:
    for col in (4, 3, 2):
        value = ws.cell(row=row, column=col).value
        if value not in (None, ""):
            return str(value)
    return ""


def find_row(
    ws,
    label: str,
    section_after: str | None,
    section_before: str | None,
    occurrence: int = 1,
) -> int:
    label_norm = normalize(label)
    after_row = 1
    before_row = ws.max_row + 1

    if section_after:
        section_norm = normalize(section_after)
        for row in range(1, ws.max_row + 1):
            if normalize(row_label(ws, row)) == section_norm:
                after_row = row
                break

    if section_before:
        section_norm = normalize(section_before)
        for row in range(after_row + 1, ws.max_row + 1):
            if normalize(row_label(ws, row)) == section_norm:
                before_row = row
                break

    matches = [
        row
        for row in range(after_row + 1, before_row)
        if normalize(row_label(ws, row)) == label_norm
    ]
    if not matches:
        scope = f" after {section_after!r}" if section_after else ""
        raise ValueError(f"Could not find row label {label!r}{scope}.")
    if occurrence == -1:
        return matches[-1]
    if occurrence < 1:
        raise ValueError(f"Invalid occurrence {occurrence} for row label {label!r}.")
    if occurrence == 1 and len(matches) > 1:
        raise ValueError(f"Found multiple rows for label {label!r}: {matches}")
    if len(matches) < occurrence:
        raise ValueError(
            f"Could not find occurrence {occurrence} for row label {label!r}; matches: {matches}"
        )
    return matches[occurrence - 1]


def find_confidence_row(ws, label: str) -> int:
    return find_row(ws, label, "2. CONFIDENCE LEVEL", "Inputs and calculations")


def scaled_excel_value(raw_value: Decimal, mapping: WorkbookMapping) -> float | int | None:
    value = abs(raw_value) if mapping.absolute else raw_value
    if mapping.blank_if_zero and value == 0:
        return None
    if mapping.scale == 1:
        return int(value) if value == value.to_integral_value() else float(value)
    return float(value / mapping.scale)


def source_comment(report: EstonianReport, year: int, item: str) -> Comment:
    source = report.get_source(year, item) or "Annual report"
    return Comment(f"Gain:\nAR{year}\n{source}", "Codex")


def cell_is_empty(cell) -> bool:
    return cell.value in (None, "")


def set_formula_font(cell) -> None:
    font = copy(cell.font)
    font.color = "000000"
    cell.font = font


def format_column_runs(columns: Iterable[int]) -> str:
    sorted_columns = sorted(set(columns))
    if not sorted_columns:
        return "none"

    ranges: list[str] = []
    start = previous = sorted_columns[0]
    for column in sorted_columns[1:]:
        if column == previous + 1:
            previous = column
            continue
        ranges.append(column_range_label(start, previous))
        start = previous = column
    ranges.append(column_range_label(start, previous))
    return ", ".join(ranges)


def column_range_label(start: int, end: int) -> str:
    start_letter = get_column_letter(start)
    end_letter = get_column_letter(end)
    return start_letter if start == end else f"{start_letter}:{end_letter}"


def group_unused_year_columns(ws, jobs: list[FillJob], dry_run: bool) -> list[str]:
    used_years = {job.year for job in jobs}
    year_columns = [
        (cell.column, year)
        for cell in ws[HEADER_ROW]
        for year in [header_year(cell.value)]
        if year is not None
    ]
    if not year_columns:
        return ["SKIP year column grouping: no year columns found"]

    used_columns: list[int] = []
    unused_columns: list[int] = []
    for column, year in year_columns:
        dimension = ws.column_dimensions[get_column_letter(column)]
        if year in used_years:
            used_columns.append(column)
            if not dry_run:
                dimension.hidden = False
                dimension.outlineLevel = 0
                dimension.collapsed = False
        else:
            unused_columns.append(column)
            if not dry_run:
                dimension.hidden = True
                dimension.outlineLevel = 1
                dimension.collapsed = False
    return [
        f"SHOW year columns: {format_column_runs(used_columns)}",
        f"GROUP hidden unused year columns: {format_column_runs(unused_columns)}",
    ]


def find_cagr_column(ws) -> int:
    for cell in ws[HEADER_ROW]:
        if isinstance(cell.value, str) and normalize(cell.value).startswith("cagr"):
            return cell.column
    raise ValueError(f"Could not find the CAGR column in row {HEADER_ROW}.")


def update_cagr_formulas(ws, jobs: list[FillJob], dry_run: bool) -> list[str]:
    years = sorted({job.year for job in jobs})
    if len(years) < 2:
        return ["SKIP CAGR formulas: fewer than two years planned"]

    cagr_years = years[-5:]
    start_year = cagr_years[0]
    end_year = cagr_years[-1]
    start_col = find_year_column(ws, start_year)
    end_col = find_year_column(ws, end_year)
    cagr_column = find_cagr_column(ws)
    start_letter = get_column_letter(start_col)
    end_letter = get_column_letter(end_col)
    cagr_letter = get_column_letter(cagr_column)

    updated = 0
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=cagr_column)
        if not (isinstance(cell.value, str) and cell.value.startswith("=")):
            continue
        formula = (
            f'=IFERROR(({end_letter}{row}/{start_letter}{row})^'
            f'(1/({end_letter}${HEADER_ROW}-{start_letter}${HEADER_ROW}))-1,"")'
        )
        if not dry_run:
            cell.value = formula
            set_formula_font(cell)
        updated += 1

    header_cell = ws.cell(row=HEADER_ROW, column=cagr_column)
    if isinstance(header_cell.value, str) and normalize(header_cell.value).startswith("cagr"):
        if not dry_run:
            header_cell.value = f"CAGR {start_year}-{end_year}"

    return [
        f"LINK {cagr_letter} CAGR formulas: {start_year}-{end_year} "
        f"({start_letter}:{end_letter}) across {updated} existing formula cells"
    ]


def combined_report_for_year(
    year: int,
    reports_desc: list[EstonianReport],
    preferred_report: EstonianReport,
) -> EstonianReport:
    source_year = preferred_report.year
    combined = EstonianReport(
        Path(f"combined_FY{year}"),
        f"FY{year} sourced primarily from AR{source_year}",
        [],
        preferred_report.terms,
        period_start=preferred_report.period_start,
        period_end=preferred_report.period_end,
        company=preferred_report.company,
        accounting_basis=preferred_report.accounting_basis,
    )

    ordered_reports = [preferred_report, *[report for report in reports_desc if report is not preferred_report]]
    for report in ordered_reports:
        for item, value in report.values.get(year, {}).items():
            if combined.get_value(year, item) is not None:
                continue
            combined.set_value(year, item, value, report.get_source(year, item) or "Annual report")
        for segment_by, records in report.segments.get(year, {}).items():
            combined_records = combined.segments.setdefault(year, {}).setdefault(segment_by, {})
            for label, record in records.items():
                if label not in combined_records:
                    combined_records[label] = copy(record)

    return combined


def fill_job_for_year(
    year: int,
    reports_desc: list[EstonianReport],
    preferred_report: EstonianReport,
    current_report: EstonianReport | None,
) -> FillJob:
    """Keep revised values and target-year period metadata as separate concepts."""
    period_report = current_report if current_report is not None and current_report.year == year else None
    return FillJob(
        combined_report_for_year(year, reports_desc, preferred_report),
        year,
        year,
        source_year=preferred_report.year,
        period_start=period_report.period_start if period_report else None,
        period_end=period_report.period_end if period_report else None,
    )


def build_fill_jobs(
    reports: list[EstonianReport],
    *,
    years: int,
    fill_comparative: bool,
    target_years: list[int] | None = None,
) -> tuple[list[FillJob], list[str]]:
    if years < 1:
        raise ValueError("--years must be at least 1.")
    if not reports:
        raise ValueError("At least one BDOC/PDF report is required.")

    messages: list[str] = []
    unique_by_year: dict[int, EstonianReport] = {}
    for report in sorted(reports, key=lambda item: (item.year, item.source_path.name), reverse=True):
        if report.year in unique_by_year:
            messages.append(
                f"SKIP duplicate FY{report.year} report: {report.source_path.name} "
                f"(using {unique_by_year[report.year].source_path.name})"
            )
            continue
        unique_by_year[report.year] = report

    sorted_reports = sorted(unique_by_year.values(), key=lambda item: item.year, reverse=True)

    if target_years is not None:
        requested_years = sorted(set(target_years), reverse=True)[:years]
        jobs: list[FillJob] = []
        for year in requested_years:
            later_report = unique_by_year.get(year + 1)
            current_report = unique_by_year.get(year)
            preferred_report = (
                later_report
                if later_report is not None
                and (later_report.values.get(year) or later_report.segments.get(year))
                else current_report
            )
            if preferred_report is None:
                preferred_report = next(
                    (
                        report
                        for report in sorted_reports
                        if report.values.get(year) or report.segments.get(year)
                    ),
                    None,
                )
            if preferred_report is None:
                messages.append(f"MISS FY{year}: no current or comparative report values")
                continue
            jobs.append(fill_job_for_year(year, sorted_reports, preferred_report, current_report))
        return jobs, messages

    target_years: list[int] = []
    seen_years: set[int] = set()

    if len(sorted_reports) == 1:
        report = sorted_reports[0]
        jobs = [
            FillJob(
                report,
                report.year,
                report.year,
                source_year=report.year,
                period_start=report.period_start,
                period_end=report.period_end,
            )
        ]
        seen_years.add(report.year)
        if fill_comparative:
            jobs.append(
                FillJob(
                    report,
                    report.year - 1,
                    report.year - 1,
                    source_year=report.year,
                )
            )
        return jobs[:years], messages

    for report in sorted_reports:
        if report.year in seen_years:
            continue
        target_years.append(report.year)
        seen_years.add(report.year)
        if len(target_years) >= years:
            break

    if len(target_years) < years and fill_comparative:
        for report in sorted_reports:
            comparative_year = report.year - 1
            if comparative_year in seen_years:
                continue
            target_years.append(comparative_year)
            seen_years.add(comparative_year)
            if len(target_years) >= years:
                break

    jobs = []
    for year in target_years[:years]:
        later_report = unique_by_year.get(year + 1)
        current_report = unique_by_year.get(year)
        preferred_report = (
            later_report
            if later_report is not None
            and (later_report.values.get(year) or later_report.segments.get(year))
            else current_report
        )
        if preferred_report is None:
            preferred_report = next(
                report
                for report in sorted_reports
                if report.values.get(year) or report.segments.get(year)
            )
        jobs.append(fill_job_for_year(year, sorted_reports, preferred_report, current_report))
    return jobs[:years], messages


def segment_value(records: dict[str, SegmentRecord], label: str) -> Decimal:
    record = records.get(label)
    return record.value if record is not None else Decimal("0")


def segment_source(records: dict[str, SegmentRecord], label: str) -> str | None:
    record = records.get(label)
    return record.source if record is not None else None


def rest_record_uses_group_total(
    first_records: dict[str, SegmentRecord],
    last_records: dict[str, SegmentRecord],
    label: str,
) -> bool:
    source = segment_source(first_records, label) or segment_source(last_records, label) or ""
    return "kokku" in normalize(source)


def segment_sort_key(
    label: str,
    first_records: dict[str, SegmentRecord],
    last_records: dict[str, SegmentRecord],
) -> tuple[int, int, str]:
    record = last_records.get(label) or first_records.get(label)
    return (0 if (record and record.group == "EU") else 1, record.order if record else 999, label)


def build_segment_summary_rows(
    first_records: dict[str, SegmentRecord],
    last_records: dict[str, SegmentRecord],
) -> list[SegmentSummaryRow]:
    labels = sorted(
        (
            label
            for label in set(first_records) | set(last_records)
            if segment_value(first_records, label) != 0 or segment_value(last_records, label) != 0
        ),
        key=lambda label: segment_sort_key(label, first_records, last_records),
    )
    if not labels:
        return []

    rest_labels = {
        label
        for label in labels
        if (first_records.get(label) and first_records[label].is_rest)
        or (last_records.get(label) and last_records[label].is_rest)
    }
    all_explicit_labels = [label for label in labels if label not in rest_labels]
    collapse_to_rest_groups = {
        "EU" if label == "Rest of EU" else "WORLD" if label == "Rest of the world" else ""
        for label in rest_labels
        if rest_record_uses_group_total(first_records, last_records, label)
    }
    force_rest_labels = {
        label
        for label in all_explicit_labels
        if (last_records.get(label) or first_records.get(label))
        and (last_records.get(label) or first_records.get(label)).group in collapse_to_rest_groups
        and not (
            (last_records.get(label) or first_records.get(label)).group == "EU"
            and label == "Estonia"
        )
    }
    eu_explicit_labels = [
        label
        for label in all_explicit_labels
        if (last_records.get(label) or first_records.get(label))
        and (last_records.get(label) or first_records.get(label)).group == "EU"
    ]
    if "Rest of EU" in rest_labels and len(eu_explicit_labels) > 8:
        force_rest_labels.update(
            label
            for label in eu_explicit_labels
            if segment_value(first_records, label) < Decimal("300000")
            or segment_value(last_records, label) < Decimal("1000000")
        )
    if "Rest of the world" in rest_labels:
        force_rest_labels.update(
            label
            for label in all_explicit_labels
            if label in {"Australia", "Hong Kong"}
            and segment_value(last_records, label) < Decimal("1000000")
        )
        force_rest_labels.update(
            label
            for label in all_explicit_labels
            if (last_records.get(label) or first_records.get(label))
            and (last_records.get(label) or first_records.get(label)).group == "WORLD"
            and segment_value(last_records, label) == 0
        )
    small_explicit_labels = {
        label
        for label in all_explicit_labels
        if (last_records.get(label) or first_records.get(label))
        and (last_records.get(label) or first_records.get(label)).group in {"EU", "WORLD"}
        and max(abs(segment_value(first_records, label)), abs(segment_value(last_records, label))) < MIN_SEGMENT_DISPLAY_VALUE
    }
    for label in small_explicit_labels:
        record = last_records.get(label) or first_records.get(label)
        if record and record.group == "EU":
            rest_labels.add("Rest of EU")
        elif record and record.group == "WORLD":
            rest_labels.add("Rest of the world")
    explicit_labels = [
        label
        for label in all_explicit_labels
        if label not in small_explicit_labels and label not in force_rest_labels
    ]

    capacity = SEGMENT_ROW_END - SEGMENT_ROW_START + 1
    selected_explicit = list(explicit_labels)
    if len(labels) > capacity:
        reserve_groups: set[str] = set()
        for label in rest_labels:
            record = last_records.get(label) or first_records.get(label)
            group = (
                "EU"
                if label == "Rest of EU"
                else "WORLD"
                if label == "Rest of the world"
                else record.group
                if record is not None
                else ""
            )
            if group:
                reserve_groups.add(group)
        slots = max(0, capacity - max(1, len(reserve_groups)))
        ranked_explicit = sorted(
            explicit_labels,
            key=lambda label: max(abs(segment_value(first_records, label)), abs(segment_value(last_records, label))),
            reverse=True,
        )
        selected_set = set(ranked_explicit[:slots])
        dropped = [label for label in explicit_labels if label not in selected_set]
        for label in dropped:
            record = last_records.get(label) or first_records.get(label)
            if record and record.group:
                reserve_groups.add(record.group)
        slots = max(0, capacity - len(reserve_groups))
        selected_set = set(ranked_explicit[:slots])
        selected_explicit = [
            label
            for label in explicit_labels
            if label in selected_set
        ]
        rest_labels = {
            "Rest of EU" if group == "EU" else "Rest of the world" if group == "WORLD" else "Other"
            for group in reserve_groups
        }

    rows_by_label: dict[str, SegmentSummaryRow] = {}
    for label in selected_explicit:
        rows_by_label[label] = SegmentSummaryRow(
            label,
            segment_value(first_records, label),
            segment_value(last_records, label),
            segment_source(first_records, label),
            segment_source(last_records, label),
        )

    for rest_label in rest_labels:
        group = "EU" if rest_label == "Rest of EU" else "WORLD" if rest_label == "Rest of the world" else ""
        first_value = segment_value(first_records, rest_label)
        last_value = segment_value(last_records, rest_label)
        for label in all_explicit_labels:
            if label in rows_by_label:
                continue
            record = last_records.get(label) or first_records.get(label)
            if record and record.group == group:
                first_value += segment_value(first_records, label)
                last_value += segment_value(last_records, label)
        rows_by_label[rest_label] = SegmentSummaryRow(
            rest_label,
            first_value,
            last_value,
            segment_source(first_records, rest_label) or "Revenue note geography: aggregated smaller countries",
            segment_source(last_records, rest_label) or "Revenue note geography: aggregated smaller countries",
        )

    return [
        rows_by_label[label]
        for label in sorted(rows_by_label, key=lambda label: segment_sort_key(label, first_records, last_records))
    ][:capacity]


def segment_rows_tie_to_revenue(
    rows: list[SegmentSummaryRow],
    first_revenue: Decimal | None,
    last_revenue: Decimal | None,
) -> bool:
    tolerance = Decimal("1")
    first_sum = sum(row.first_value for row in rows)
    last_sum = sum(row.last_value for row in rows)
    if first_revenue is not None and abs(first_sum - first_revenue) > tolerance:
        return False
    if last_revenue is not None and abs(last_sum - last_revenue) > tolerance:
        return False
    return True


def segment_comment(year: int, source: str | None) -> Comment:
    return Comment(f"Gain:\nAR{year}\n{source or 'Revenue note geography'}", "Codex")


def segment_excel_value(value: Decimal) -> float | int:
    scaled = value / EUR_TO_EURM
    return int(scaled) if scaled == scaled.to_integral_value() else float(scaled)


def build_revenue_segment_plans(jobs: list[FillJob]) -> list[SegmentPlan]:
    sorted_jobs = sorted(jobs, key=lambda job: job.year)
    first_job = sorted_jobs[0]
    last_job = sorted_jobs[-1]
    first_year = first_job.value_year
    last_year = last_job.value_year

    plans_by_segment: dict[str, SegmentPlan] = {}
    for segment_by in (GEOGRAPHY_SEGMENT, ACTIVITY_SEGMENT):
        first_records = first_job.report.get_segments(first_year, segment_by)
        last_records = last_job.report.get_segments(last_year, segment_by)
        rows = build_segment_summary_rows(first_records, last_records)
        if rows:
            if segment_by == ACTIVITY_SEGMENT:
                if len(rows) <= 1 or len(rows) > 8:
                    continue
                if not segment_rows_tie_to_revenue(
                    rows,
                    first_job.report.get_value(first_year, "Revenue"),
                    last_job.report.get_value(last_year, "Revenue"),
                ):
                    continue
            plans_by_segment[segment_by] = SegmentPlan(
                "Net revenue",
                activity_segment_display_name(rows) if segment_by == ACTIVITY_SEGMENT else segment_by,
                first_year,
                last_year,
                rows,
            )

    geography = plans_by_segment.get(GEOGRAPHY_SEGMENT)
    activity = plans_by_segment.get(ACTIVITY_SEGMENT)
    if geography and activity:
        if len(activity.rows) >= len(geography.rows):
            return [activity, geography]
        return [geography, activity]
    return [plan for plan in (geography, activity) if plan is not None]


def activity_segment_display_name(rows: list[SegmentSummaryRow]) -> str:
    labels = {row.label for row in rows}
    if {"Retail", "Wholesale"} & labels:
        return "Type"
    if {"Programming & software", "Consulting", "Data processing & web hosting"} & labels:
        return "Source"
    return ACTIVITY_SEGMENT


def fill_segment_block(
    ws,
    plan: SegmentPlan,
    block: tuple[int, int, int, int],
    *,
    overwrite: bool,
    add_comments: bool,
    dry_run: bool,
) -> list[str]:
    label_col, first_value_col, last_value_col, cagr_col = block
    messages = [
        f"=== Filling Segments: {plan.segmentation_of} by {plan.segment_by} "
        f"({plan.first_year}-{plan.last_year}) ==="
    ]
    if not dry_run:
        ws.cell(row=4, column=first_value_col).value = plan.segmentation_of
        ws.cell(row=5, column=first_value_col).value = plan.segment_by
        ws.cell(row=8, column=first_value_col).value = plan.first_year
        ws.cell(row=8, column=last_value_col).value = plan.last_year

    for row_idx in range(SEGMENT_ROW_START, SEGMENT_ROW_END + 1):
        if not dry_run and overwrite:
            for col_idx in (label_col, first_value_col, last_value_col):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = None
                cell.comment = None

    first_letter = get_column_letter(first_value_col)
    last_letter = get_column_letter(last_value_col)
    for offset, segment_row in enumerate(plan.rows):
        row_idx = SEGMENT_ROW_START + offset
        label_cell = ws.cell(row=row_idx, column=label_col)
        first_cell = ws.cell(row=row_idx, column=first_value_col)
        last_cell = ws.cell(row=row_idx, column=last_value_col)
        if not overwrite and not (cell_is_empty(label_cell) and cell_is_empty(first_cell) and cell_is_empty(last_cell)):
            messages.append(f"SKIP Segments row {row_idx}: already populated")
            continue
        if not dry_run:
            label_cell.value = segment_row.label
            first_cell.value = segment_excel_value(segment_row.first_value)
            last_cell.value = segment_excel_value(segment_row.last_value)
            if add_comments:
                first_cell.comment = segment_comment(plan.first_year, segment_row.first_source)
                last_cell.comment = segment_comment(plan.last_year, segment_row.last_source)
            cagr_cell = ws.cell(row=row_idx, column=cagr_col)
            if not (isinstance(cagr_cell.value, str) and cagr_cell.value.startswith("=")):
                cagr_cell.value = (
                    f'=IFERROR(({last_letter}{row_idx}/{first_letter}{row_idx})'
                    f'^(1/({last_letter}$8-{first_letter}$8))-1,"")'
                )
        messages.append(
            f"SET  Segments {plan.segment_by} {segment_row.label}: "
            f"{segment_excel_value(segment_row.first_value)} / {segment_excel_value(segment_row.last_value)}"
        )
    return messages


def fill_segments_sheet(
    wb,
    jobs: list[FillJob],
    *,
    overwrite: bool,
    add_comments: bool,
    dry_run: bool,
) -> list[str]:
    if SEGMENTS_SHEET not in wb.sheetnames:
        return [f"SKIP Segments: workbook does not contain {SEGMENTS_SHEET!r} sheet"]
    if len(jobs) < 2:
        return ["SKIP Segments: fewer than two fiscal years planned"]

    plans = build_revenue_segment_plans(jobs)
    if not plans:
        return ["MISS Segments: no revenue segmentation data found"]

    ws = wb[SEGMENTS_SHEET]
    messages: list[str] = []
    for plan, block in zip(plans, SEGMENT_BLOCKS):
        messages.extend(
            fill_segment_block(
                ws,
                plan,
                block,
                overwrite=overwrite,
                add_comments=add_comments,
                dry_run=dry_run,
            )
        )
    return messages


def find_annualisation_row(ws, label: str) -> int:
    return find_row(ws, label, "Annualisation", None)


def set_formula_cell(ws, row: int, col: int, formula: str, dry_run: bool) -> None:
    if dry_run:
        return
    cell = ws.cell(row=row, column=col)
    cell.value = formula
    cell.comment = None
    set_formula_font(cell)


def configure_annualisation(
    ws,
    job: FillJob,
    year: int,
    col: int,
    dry_run: bool,
) -> tuple[dict[str, int], list[str]]:
    if job.period_start is None or job.period_end is None:
        raise ValueError(f"Could not identify the reporting period dates for FY{year}.")
    col_letter = get_column_letter(col)
    annualisation_start_row = find_row(ws, "Annualisation", None, None)
    annualisation_end_row = find_annualisation_row(ws, "YoY growth (%)")
    labels = (
        "Starting date",
        "Ending date",
        "Annualisation factor",
        "Footnote (copy on CMS)",
        "Net revenue",
        "Other income",
        "Revenue",
        "COGS",
        "Other COGS",
        "Gross margin",
        "Reported EBIT",
        "Reported EBIT Check",
        "Total depreciation",
        "Total amortisation",
        "D&A",
        "EBITDA",
        "Investments in tangible assets",
        "Investments in intangible assets",
        "CAPEX",
    )
    rows = {label: find_annualisation_row(ws, label) for label in labels}

    if not dry_run:
        for row in range(annualisation_start_row, annualisation_end_row + 1):
            ws.row_dimensions[row].hidden = False
        start_cell = ws.cell(row=rows["Starting date"], column=col)
        end_cell = ws.cell(row=rows["Ending date"], column=col)
        start_cell.value = job.period_start
        end_cell.value = job.period_end
        start_cell.number_format = "dd-mm-yyyy"
        end_cell.number_format = "dd-mm-yyyy"

    set_formula_cell(
        ws,
        rows["Annualisation factor"],
        col,
        (
            f'=IF(AND({col_letter}{rows["Starting date"]}<>0,'
            f'{col_letter}{rows["Ending date"]}<>0),'
            f'({col_letter}{rows["Ending date"]}-{col_letter}{rows["Starting date"]}+1)/365,"")'
        ),
        dry_run,
    )
    set_formula_cell(
        ws,
        rows["Footnote (copy on CMS)"],
        col,
        (
            f'=IF({col_letter}{rows["Annualisation factor"]}<>"","FY"&{col_letter}${HEADER_ROW}'
            f'&": the Company reported for the period from "&TEXT({col_letter}{rows["Starting date"]},"dd-mm")'
            f'&" to "&TEXT({col_letter}{rows["Ending date"]},"dd-mm")'
            f'&". Figures have been annualised to reflect the full-year effect, assuming no seasonality.","")'
        ),
        dry_run,
    )

    scratchpad_formulas = {
        "Revenue": f'=IFERROR({col_letter}{rows["Net revenue"]}+{col_letter}{rows["Other income"]},"")',
        "Gross margin": (
            f'={col_letter}{rows["Revenue"]}-IFERROR({col_letter}{rows["COGS"]}+'
            f'{col_letter}{rows["Other COGS"]},"")'
        ),
        "Reported EBIT Check": (
            f'={col_letter}{rows["EBITDA"]}-{col_letter}{rows["Total depreciation"]}'
            f'-{col_letter}{rows["Total amortisation"]}'
        ),
        "D&A": (
            f'=SUM({col_letter}{rows["Total depreciation"]}:'
            f'{col_letter}{rows["Total amortisation"]})'
        ),
        "EBITDA": f'=IFERROR({col_letter}{rows["D&A"]}+{col_letter}{rows["Reported EBIT"]},"")',
        "CAPEX": (
            f'=SUM({col_letter}{rows["Investments in tangible assets"]}:'
            f'{col_letter}{rows["Investments in intangible assets"]})'
        ),
    }
    for label, formula in scratchpad_formulas.items():
        set_formula_cell(ws, rows[label], col, formula, dry_run)

    reported_rows = {
        label: find_row(ws, label, "1. REPORTED FIGURES", "2. ADJUSTMENTS ")
        for label in ("Net revenue", "Other income", "Revenue", "COGS", "Other COGS", "Gross margin", "Reported EBIT", "D&A", "EBITDA", "CAPEX")
    }
    reported_formulas = {
        "Revenue": f'=IFERROR({col_letter}{reported_rows["Net revenue"]}+{col_letter}{reported_rows["Other income"]},"")',
        "Gross margin": (
            f'={col_letter}{reported_rows["Revenue"]}-IFERROR({col_letter}{reported_rows["COGS"]}+'
            f'{col_letter}{reported_rows["Other COGS"]},"")'
        ),
        "EBITDA": f'=IFERROR({col_letter}{reported_rows["D&A"]}+{col_letter}{reported_rows["Reported EBIT"]},"")',
    }
    for label, formula in reported_formulas.items():
        set_formula_cell(ws, reported_rows[label], col, formula, dry_run)

    return rows, [
        f"ANNUALISE FY{year}: {job.period_start.isoformat()} to {job.period_end.isoformat()} ({job.period_days} days)",
        f"SHOW annualisation rows: {annualisation_start_row}:{annualisation_end_row}",
    ]


def fill_period(
    ws,
    job: FillJob,
    *,
    overwrite: bool,
    add_comments: bool,
    fill_cogs: bool,
    fill_goodwill_amortisation: bool,
    dry_run: bool,
) -> list[str]:
    col = find_year_column(ws, job.year)
    col_letter = get_column_letter(col)
    messages: list[str] = []
    confidence_labels: set[str] = set()
    annualisation_rows: dict[str, int] | None = None
    if job.requires_annualisation:
        annualisation_rows, annualisation_messages = configure_annualisation(
            ws, job, job.year, col, dry_run
        )
        messages.extend(annualisation_messages)

    enabled_flags = {
        "fill_cogs": fill_cogs,
        "fill_goodwill_amortisation": fill_goodwill_amortisation,
    }

    for mapping in WORKBOOK_MAPPINGS:
        if mapping.optional_flag and not enabled_flags.get(mapping.optional_flag, False):
            messages.append(f"SKIP {col_letter} {mapping.row_label}: optional mapping disabled")
            continue

        row = find_row(
            ws,
            mapping.row_label,
            mapping.section_after,
            mapping.section_before,
            mapping.occurrence,
        )
        input_row = (
            annualisation_rows[mapping.row_label]
            if annualisation_rows
            and mapping.row_label
            in {
                "Net revenue",
                "Other income",
                "COGS",
                "Reported EBIT",
                "Total depreciation",
                "Total amortisation",
                "D&A",
                "Investments in tangible assets",
                "Investments in intangible assets",
                "CAPEX",
            }
            else row
        )
        cell = ws.cell(row=input_row, column=col)
        reported_cell = ws.cell(row=row, column=col)
        raw_value = job.report.get_value(job.value_year, mapping.item)

        if mapping.item in {"Total depreciation", "Total amortisation"} and job.report.get_value(
            job.value_year, "D&A"
        ) is not None:
            messages.append(f"SKIP {col_letter}{row} {mapping.row_label}: total D&A row available")
            continue

        if mapping.item in {"Investments in tangible assets", "Investments in intangible assets"} and job.report.get_value(
            job.value_year, "CAPEX"
        ) is not None:
            messages.append(f"SKIP {col_letter}{row} {mapping.row_label}: total CAPEX row available")
            continue

        if mapping.item == "D&A" and raw_value is None and (
            job.report.get_value(job.value_year, "Total depreciation") is not None
            or job.report.get_value(job.value_year, "Total amortisation") is not None
        ):
            messages.append(f"SKIP {col_letter}{row} D&A: split depreciation/amortisation rows populated")
            continue

        if mapping.item == "CAPEX" and raw_value is None and (
            job.report.get_value(job.value_year, "Investments in tangible assets") is not None
            or job.report.get_value(job.value_year, "Investments in intangible assets") is not None
        ):
            messages.append(f"SKIP {col_letter}{row} CAPEX: split tangible/intangible rows populated")
            continue

        if raw_value is None:
            if overwrite and not dry_run:
                cell.value = None
                cell.comment = None
                if input_row != row:
                    reported_cell.value = None
                    reported_cell.comment = None
            messages.append(f"MISS {col_letter}{input_row} {mapping.row_label}: no value found")
            continue

        value = scaled_excel_value(raw_value, mapping)
        if not overwrite and not cell_is_empty(cell):
            messages.append(f"SKIP {col_letter}{row} {mapping.row_label}: cell already populated")
            continue

        if mapping.confidence_label:
            confidence_labels.add(mapping.confidence_label)

        if not dry_run:
            cell.value = value
            if add_comments and value is not None:
                cell.comment = source_comment(job.report, job.value_year, mapping.item)
            if annualisation_rows and input_row != row:
                reported_cell.value = (
                    f"={col_letter}{input_row}/${col_letter}${annualisation_rows['Annualisation factor']}"
                )
                reported_cell.comment = None
                set_formula_font(reported_cell)

        rendered = "" if value is None else f"{value:.8f}".rstrip("0").rstrip(".")
        source = job.report.get_source(job.value_year, mapping.item) or "n/a"
        messages.append(f"SET  {col_letter}{input_row} {mapping.row_label}: {rendered} from {source}")
        if annualisation_rows and input_row != row:
            messages.append(
                f"LINK {col_letter}{row} {mapping.row_label}: "
                f"={col_letter}{input_row}/${col_letter}${annualisation_rows['Annualisation factor']}"
            )

    if "Adjusted EBITDA" in confidence_labels:
        confidence_labels.add("Adjusted EBIT")

    for confidence_label in sorted(confidence_labels):
        try:
            confidence_row = find_confidence_row(ws, confidence_label)
        except ValueError:
            messages.append(f"MISS {col_letter} confidence {confidence_label}: row not found")
            continue

        confidence_cell = ws.cell(row=confidence_row, column=col)
        if overwrite or cell_is_empty(confidence_cell):
            if not dry_run:
                confidence_cell.value = "Actual"
            messages.append(f"CONF {col_letter}{confidence_row} {confidence_label}: Actual")
        else:
            messages.append(f"SKIP {col_letter}{confidence_row} {confidence_label}: already populated")

    return messages


def save_workbook(wb, output_path: Path) -> None:
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except AttributeError:
        pass
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def fill_workbook(args: argparse.Namespace) -> list[str]:
    terms = load_terms(args.mapping_json)
    input_paths = [path for path_group in args.bdoc for path in path_group]
    reports = [
        report
        for path in input_paths
        for report in parse_reports_from_input(path, terms)
    ]
    jobs, messages = build_fill_jobs(
        reports,
        years=args.years,
        fill_comparative=args.fill_comparative,
    )

    wb = load_workbook(args.template)
    if FINANCIALS_SHEET not in wb.sheetnames:
        raise ValueError(f"Workbook does not contain a {FINANCIALS_SHEET!r} sheet.")
    ws = wb[FINANCIALS_SHEET]

    for job in jobs:
        messages.append(
            f"=== Filling FY{job.year} from {job.report.source_path.name} "
            f"({job.report.payload_name}) values FY{job.value_year} ==="
        )
        messages.extend(
            fill_period(
                ws,
                job,
                overwrite=not args.no_overwrite,
                add_comments=not args.no_comments,
                fill_cogs=args.fill_cogs,
                fill_goodwill_amortisation=args.fill_goodwill_amortisation,
                dry_run=args.dry_run,
            )
        )

    messages.extend(update_cagr_formulas(ws, jobs, dry_run=args.dry_run))
    messages.extend(group_unused_year_columns(ws, jobs, dry_run=args.dry_run))
    if args.fill_segments:
        messages.extend(
            fill_segments_sheet(
                wb,
                jobs,
                overwrite=not args.no_overwrite,
                add_comments=not args.no_comments,
                dry_run=args.dry_run,
            )
        )
    else:
        messages.append("SKIP Segments: optional mapping disabled")

    if not args.dry_run:
        save_workbook(wb, args.output)
        messages.append(f"Saved: {args.output}")
    return messages


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Fill a Gain.pro Financials template from Estonian BDOC annual reports."
    )
    parser.add_argument(
        "--bdoc",
        required=True,
        nargs="+",
        action="append",
        type=Path,
        help="Input .bdoc/.ddoc/.asice, PDF annual report file(s), or ZIP files containing annual-report PDFs.",
    )
    parser.add_argument("--template", required=True, type=Path, help="Input Excel template/workbook.")
    parser.add_argument("--output", required=True, type=Path, help="Output .xlsx path.")
    parser.add_argument("--mapping-json", type=Path, help="Optional Estonian item mapping JSON.")
    parser.add_argument("--years", type=int, default=6, help="Maximum number of fiscal years to fill. Default: 6.")
    parser.add_argument(
        "--fill-comparative",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="For one report, also fill the previous-year comparative column. Default: true.",
    )
    parser.add_argument(
        "--fill-cogs",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="Fill COGS and Gross margin confidence rows when available. Default: true.",
    )
    parser.add_argument(
        "--fill-goodwill-amortisation",
        action="store_true",
        help="Fill goodwill amortisation in the adjustments section when disclosed.",
    )
    parser.add_argument(
        "--fill-segments",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="Fill the Segments sheet with revenue segmentations when available. Default: true.",
    )
    parser.add_argument("--no-overwrite", action="store_true", help="Do not overwrite existing values in mapped cells.")
    parser.add_argument("--no-comments", action="store_true", help="Do not add annual-report source comments.")
    parser.add_argument("--dry-run", action="store_true", help="Print planned changes without saving.")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv or sys.argv[1:])
    try:
        messages = fill_workbook(args)
    except Exception as exc:
        raise SystemExit(str(exc)) from exc

    for message in messages:
        print(message)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
